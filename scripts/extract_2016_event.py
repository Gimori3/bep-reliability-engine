"""Extract the 2016 typhoon event observations into compact processed CSVs.

One-time, re-runnable extraction from the raw agency data drop
``data/digitized/2016_event_data/`` (188 MB, gitignored) into small committed
CSVs under ``data/processed/2016_event/`` that the Phase 2 observed-event
loader (``bayesian_reliability_updating.events``) consumes. Run from the
repository root::

    python scripts/extract_2016_event.py

Sources (see ``data/processed/2016_event/README.md`` for the full audit):

* ``36_2016.8.20-8.31/観測所水位・流量データ/H_Q_2016.8.20-8.31.xlsx``:
  hourly observed stage [m MSL] and published discharge [m^3/s] at 30
  stations across the Tokachi river system, August and September 2016.
  Only the two study rivers (Tokachi, Satsunai) are extracted.
  KNOWN SOURCE ANOMALY: the sheet ``時刻水位201609`` (September stage) is a
  corrupted duplicate of the September discharge sheet (its header row and
  values are discharge); September stage is therefore NOT extractable and is
  not emitted. The full erosive window of the event is contained in the
  August sheet (checked in the loader; the recession is below every study
  section's landside toe by the window end).
* ``36_2016.8.20-8.31/洪水痕跡水位/H28_kon.xlsx``: the September 2016
  post-flood trace survey (洪水痕跡標高一覧表): per-KP left/right levee
  trace elevations [m MSL] plus the design high water level, at 0.2 km
  spacing. Extracted for the two study rivers.

Everything numeric is emitted verbatim (no smoothing, no gap filling); the
non-numeric sentinels of the source ('閉局' station closed, '欠測' missing,
'-') become empty cells.
"""

from __future__ import annotations

import argparse
import csv
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = REPO_ROOT / "data" / "digitized" / "2016_event_data"
HQ_WORKBOOK = (
    RAW_DIR / "36_2016.8.20-8.31" / "観測所水位・流量データ" / "H_Q_2016.8.20-8.31.xlsx"
)
TRACE_WORKBOOK = RAW_DIR / "36_2016.8.20-8.31" / "洪水痕跡水位" / "H28_kon.xlsx"
OUT_DIR = REPO_ROOT / "data" / "processed" / "2016_event"

# The study rivers (the only rivers with production cross-sections).
STUDY_RIVERS: tuple[str, ...] = ("Tokachi", "Satsunai")
RIVER_JA: dict[str, str] = {"Tokachi": "十勝川", "Satsunai": "札内川"}

# Station romanization: the processed CSVs carry ASCII column names so no
# downstream consumer depends on CJK-safe I/O. Order follows the workbook.
STATION_ASCII: dict[str, str] = {
    "共栄橋": "kyoeibashi",
    "熊牛": "kumaushi",
    "芽室太": "memurobuto",
    "帯広": "obihiro",
    "十勝中央大橋": "tokachi_chuo_ohashi",
    "千代田": "chiyoda",
    "茂岩": "moiwa",
    "大津": "otsu",
    "竜潭上流": "ryutan_joryu",
    "札内ダム直下": "satsunai_dam_chokka",
    "南札内": "minami_satsunai",
    "上札内": "kami_satsunai",
    "第二大川橋": "daini_okawabashi",
    "南帯橋": "nantaibashi",
    "札内": "satsunai",
}

# Sheet name -> (quantity, month tag). The September stage sheet is EXCLUDED:
# it is a corrupted duplicate of the September discharge sheet (see module
# docstring); emitting it as stage would be a silent unit error.
HQ_SHEETS: dict[str, tuple[str, str]] = {
    "時刻水位201608": ("stage", "201608"),
    "時刻流量201608": ("discharge", "201608"),
    "時刻流量201609": ("discharge", "201609"),
}


def _numeric_or_blank(value: object) -> str:
    """Format a workbook cell: numeric verbatim, any sentinel becomes empty."""
    if isinstance(value, (int, float)):
        return repr(float(value))
    return ""


def extract_hq_sheet(sheet, river_ja: str) -> tuple[list[str], list[list[str]]]:
    """Extract one river's station columns from one H/Q sheet.

    The sheet layout is column pairs (datetime, value) with a three-row
    header: row 1 river name, row 2 station name, row 3 quantity label.
    All pairs of one sheet share the same time axis; this is asserted.

    Parameters
    ----------
    sheet : openpyxl worksheet
        One of the ``時刻水位``/``時刻流量`` sheets.
    river_ja : str
        The Japanese river name to select columns for.

    Returns
    -------
    tuple of (list of str, list of list of str)
        ``(header, rows)``: the CSV header (``datetime_jst`` plus one ASCII
        station column per selected station) and the data rows.
    """
    grid = list(sheet.iter_rows(values_only=True))
    rivers, stations = grid[0], grid[1]
    columns: list[tuple[int, str]] = []
    for c in range(1, len(stations), 2):
        if rivers[c] == river_ja and stations[c] is not None:
            station = str(stations[c]).strip()
            columns.append((c, STATION_ASCII[station]))
    if not columns:
        raise ValueError(f"no stations found for river {river_ja!r}")

    header = ["datetime_jst"] + [name for _, name in columns]
    rows: list[list[str]] = []
    for raw in grid[3:]:
        stamp = raw[columns[0][0] - 1]
        if not isinstance(stamp, datetime):
            continue
        for c, _ in columns:
            other = raw[c - 1]
            if isinstance(other, datetime) and other != stamp:
                raise ValueError(f"misaligned time axes at {stamp} vs {other}")
        rows.append(
            [stamp.strftime("%Y-%m-%dT%H:%M")]
            + [_numeric_or_blank(raw[c]) for c, _ in columns]
        )
    return header, rows


def extract_traces(workbook, river: str) -> list[list[str]]:
    """Extract one river's flood-trace table rows.

    Layout per sheet: title rows, then per-KP rows with columns KP, design
    HWL, left trace elevation, left levee name, right trace elevation,
    right levee name. Elevations are m MSL (標高), surveyed September 2016.
    """
    sheet = workbook[RIVER_JA[river]]
    rows: list[list[str]] = []
    for raw in sheet.iter_rows(min_row=6, max_col=6, values_only=True):
        kp = raw[0]
        if not isinstance(kp, (int, float)):
            continue
        rows.append(
            [
                river,
                repr(round(float(kp), 1)),
                _numeric_or_blank(raw[1]),
                _numeric_or_blank(raw[2]),
                str(raw[3] or "").strip().replace("\n", " "),
                _numeric_or_blank(raw[4]),
                str(raw[5] or "").strip().replace("\n", " "),
            ]
        )
    return rows


def _write_csv(path: Path, header: list[str], rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as stream:
        writer = csv.writer(stream)
        writer.writerow(header)
        writer.writerows(rows)
    print(f"wrote {path.relative_to(REPO_ROOT)} ({len(rows)} rows)")


def main() -> None:
    # This driver takes no arguments. The parser exists so that a probe
    # (--help, a stray flag) is inert instead of running the whole study.
    argparse.ArgumentParser(description=__doc__.splitlines()[0]).parse_args()

    hq = load_workbook(HQ_WORKBOOK, read_only=True, data_only=True)
    try:
        for sheet_name, (quantity, month) in HQ_SHEETS.items():
            for river in STUDY_RIVERS:
                header, rows = extract_hq_sheet(hq[sheet_name], RIVER_JA[river])
                _write_csv(
                    OUT_DIR / f"{quantity}_hourly_{river}_{month}.csv", header, rows
                )
    finally:
        hq.close()

    traces = load_workbook(TRACE_WORKBOOK, read_only=True, data_only=True)
    try:
        trace_rows: list[list[str]] = []
        for river in STUDY_RIVERS:
            trace_rows.extend(extract_traces(traces, river))
    finally:
        traces.close()
    _write_csv(
        OUT_DIR / "flood_trace_2016.csv",
        [
            "river",
            "kp",
            "design_hwl_m_msl",
            "trace_left_m_msl",
            "levee_left",
            "trace_right_m_msl",
            "levee_right",
        ],
        trace_rows,
    )


if __name__ == "__main__":
    main()
