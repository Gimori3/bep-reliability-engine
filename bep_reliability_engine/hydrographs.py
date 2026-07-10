"""M3 ``hydrograph_loader``: d4PDF discharge ensemble I/O and stage translation.

Single responsibility (spec §1, M3): ingest the d4PDF discharge ensemble
provided by Uemura (Docon), apply the per-KP H-Q rating curve to turn discharge
Q(t) into river stage h(t), and expose each ensemble member as a clean
:class:`HydrographRecord` (spec §2) carrying the time axis, stage series, peak,
duration, scenario tag, native resolution, and provenance. This module also
**isolates all unit handling at the ingest boundary** (spec §1,
docs/conventions.md): everything downstream of :func:`build_hydrograph_record`
works in strict SI base units.

The authoritative record for the dataset format and the conversion is
**ADR-0019** (``docs/decisions/0019-m3-hydrograph-data-and-hq-conversion.md``);
the facts below trace to it and are locked by ``tests/test_hydrographs.py``.

H-Q conversion (Uemura thesis Eq. 4.19; ADR-0019 §2-§4)
--------------------------------------------------------
::

    h_t = sqrt(Q_t / a_kp) - b_kp

with stage h in **m MSL** (the same datum as the 2019 bank-height HWL values,
so converted stages compare to HWL with no reconciliation), discharge Q in
m^3/s, and per-KP coefficients ``a_kp``/``b_kp`` from non-uniform flow
computation. ``-b_kp`` acts as an additive datum term (b runs from ~-29 at
KP 53.8 to ~-41.7 at KP 62.8 on the Tokachi). This is **NOT** a power law
``a*Q**b`` — that reading of the email's one-line summary is wrong by ~120
orders of magnitude. Validation anchor: at Obihiro (Tokachi KP 56.6,
a = 140.33, b = -32.49) a peak of ~4,180 m^3/s reproduces the 2016 record peak
of ~37.95 m MSL, 0.19 m below the 38.14 m MSL design HWL.

Stage is deliberately **not** capped at the levee crest here (ADR-0019 §8):
crest capping is a property of Uemura's overflow mechanism, not of the seepage
boundary condition the BEP engine consumes.

Layering: pure logic versus thin file readers (ADR-0019 Consequences)
----------------------------------------------------------------------
Pure, file-free, unit-boundary layer (fully covered by the tests):

* :func:`apply_rating_curve` — elementwise Eq. 4.19 with the non-negative
  radicand guard.
* :func:`build_hydrograph_record` — record construction: hours -> SI seconds,
  ``native_dt`` derived from the actual spacing, axis validation, provenance.
* :func:`parse_member_header` — ``HPB_mXXX_YYYY`` / ``HFB_{SST}_mXXX_YYYY``
  column headers to provenance (HPB -> ``'historical'``, HFB -> ``'+4K'``).
* :func:`resolve_discharge_source_kp` — the ADR-0019 §7 upper-Tokachi proxy
  rule (KP 62.0-62.8 take the KP 61.8 discharge under their own local rating).

Thin file readers (dumb parsers, no unit conversion, no physics):

* :func:`load_rating_coefficients` — the ``HQrelation_*Riv_2017.csv`` files,
  which are **Shift-JIS encoded with full-width ``HQ_a``/``HQ_b`` header
  cells** (ADR-0019 §5); the encoding is specified explicitly.
* :func:`read_discharge_ensemble` — one per-river/KP-band Excel workbook
  (sheet ``QT``, ``Time`` column in integer hours 1..192, one column per
  ensemble member), read with openpyxl in **read-only mode** because the
  workbooks carry 3,000 (HPB) / 5,400 (HFB) member columns.

:func:`load_hydrograph_ensemble` composes the two layers into the spec §2
``dict[event_id -> HydrographRecord]``. The 2016 observed hydrograph enters
through the same :class:`HydrographRecord` interface (empty provenance).

References
----------
ADR-0019 (authoritative M3 facts). Spec §1 (M3 responsibility, unit
isolation), §2 (``HydrographRecord`` contract), §11 (native resolution /
rising-limb check). docs/conventions.md (strict SI base units; unit
conversions only at the M1/M3 boundary).
"""

from __future__ import annotations

import csv
import math
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Mapping

import numpy as np
from numpy.typing import ArrayLike, NDArray

__all__ = [
    "CanonicalShape",
    "HydrographRecord",
    "apply_rating_curve",
    "build_hydrograph_record",
    "conditioning_record_for_level",
    "experiment_for_scenario",
    "flood_timescales",
    "load_canonical_shape",
    "load_hydrograph_ensemble",
    "load_rating_coefficients",
    "normalize_stage_shape",
    "parse_member_header",
    "rating_curve_path",
    "read_discharge_ensemble",
    "resolve_band_workbook",
    "resolve_discharge_source_kp",
    "validate_datum_consistency",
]

# Exact SI conversion for the ingest boundary; the source expresses time in hours.
_SECONDS_PER_HOUR: float = 3600.0

# Relative tolerance for the uniform-spacing check. The source is nominally at a
# fixed native resolution, so the spacing must be constant to within float noise;
# a genuinely irregular axis (native_dt ill-defined) is rejected.
_SPACING_RTOL: float = 1e-6

# d4PDF workbook layout (ADR-0019 §1): one sheet per file, first column Time.
_QT_SHEET_NAME: str = "QT"
_TIME_COLUMN_HEADER: str = "Time"

# Member-column header forms (ADR-0019 §1) and the scenario mapping (§9). The
# six SST patterns are the prescribed sea-surface-temperature ensemble of the
# +4K experiment, 15 members each; the past experiment is 50 flat members.
_SST_PATTERNS: tuple[str, ...] = ("CC", "GF", "HA", "MI", "MP", "MR")
_HPB_HEADER_RE = re.compile(r"HPB_(m\d+)_(\d{4})\Z")
_HFB_HEADER_RE = re.compile(rf"HFB_({'|'.join(_SST_PATTERNS)})_(m\d+)_(\d{{4}})\Z")
_SCENARIO_HISTORICAL: str = "historical"
_SCENARIO_PLUS_4K: str = "+4K"

# Rating-coefficient CSV header names after NFKC normalization (the on-disk
# header cells contain FULL-WIDTH a/b, U+FF41/U+FF42; ADR-0019 §5).
_RATING_KP_HEADER: str = "KP"
_RATING_A_HEADER: str = "HQ_a"
_RATING_B_HEADER: str = "HQ_b"

# Upper-Tokachi discharge coverage gap (ADR-0019 §7): hydrograph files stop at
# KP 61.8 while rating coefficients extend to KP 62.8, so these five nodes
# proxy the KP 61.8 discharge series under their OWN local rating.
_PROXY_SOURCE_KP: float = 61.8
_PROXY_NODE_KPS: tuple[float, ...] = (62.0, 62.2, 62.4, 62.6, 62.8)
_KP_MATCH_ATOL: float = 1e-6

# The RETIRED ADR-0018 exit-point placeholder: generated configs now carry the
# ADR-0021 landside-toe elevations in m MSL, but a hand-built or legacy config
# could still pair an MSL stage record with a zero z_toe — a datum error the
# guard below refuses (M3 stages are m MSL, ADR-0019 §3).
_PROVISIONAL_Z_TOE_M: float = 0.0

# Scenario tag -> d4PDF experiment (ADR-0019 §9, ADR-0020 §3). The single
# source of this mapping: the orchestrator selects workbooks by config.scenario
# through here, and no config field duplicates it.
_EXPERIMENT_BY_SCENARIO: dict[str, str] = {
    _SCENARIO_HISTORICAL: "HPB",
    _SCENARIO_PLUS_4K: "HFB",
}

# The two rivers the data drop covers; the rating filename derives from this
# name (ADR-0020 §2), so it is validated rather than interpolated blindly.
_KNOWN_RIVERS: tuple[str, ...] = ("Tokachi", "Satsunai")

# Band-workbook filename grammar (ADR-0020 §4): the provider's names encode
# experiment, river and the covered KP range, e.g.
# "Hydro Data, HPB, Tokachi Riv. KP056.20-KP061.80.xlsx". Parsed, not
# hardcoded, so a future data drop with new bands needs no code change.
_BAND_WORKBOOK_RE = re.compile(
    r"Hydro Data, (?P<experiment>HPB|HFB), (?P<river>Tokachi|Satsunai) Riv\. "
    r"KP(?P<lo>\d{3}\.\d{2})-KP(?P<hi>\d{3}\.\d{2})\.xlsx\Z"
)


@dataclass(frozen=True)
class HydrographRecord:
    """One loaded event: stage series plus metadata (the spec §2 M3 output).

    Emitted by :func:`build_hydrograph_record` and, per ensemble member, by
    :func:`load_hydrograph_ensemble`; both ``run.py`` hydrograph paths (the
    canonical d4PDF scaler and the synthetic stub) construct this concrete
    type. M8 reads only ``h``, ``peak`` and ``native_dt`` (duck-typed,
    ADR-0010), while the remaining fields carry provenance and the
    static-comparison scalar.

    Attributes
    ----------
    t : numpy.ndarray, shape (T,)
        Time axis in **seconds** (SI), strictly increasing and uniformly spaced.
    h : numpy.ndarray, shape (T,)
        River stage h(t) [m MSL], Eq. 4.19 applied to Q(t). Deliberately NOT
        capped at the levee crest (ADR-0019 §8).
    peak : float
        Peak stage max(h) [m MSL]; the scalar the static branch compares
        against (spec §3 step 4).
    duration_hours : float
        Elapsed event span in hours (``(t[-1] - t[0]) / 3600``).
    scenario : str
        Climate scenario tag, ``'historical'`` (HPB) or ``'+4K'`` (HFB); flows
        through to the ``FragilityResult`` metadata and the climate comparison.
    event_id : str
        Event identifier; for d4PDF members, the verbatim column header.
    native_dt : float
        Native temporal resolution [s], derived from the time-array spacing
        (hourly source => 3600.0); the default integration timestep and the
        spec §11 rising-limb-resolution check.
    provenance : dict
        ADR-0019 provenance: SST pattern where applicable, member ID, year,
        KP, and ``discharge_proxied_from`` where applicable. Empty (``{}``)
        for records without ensemble provenance, e.g. the 2016 observed
        hydrograph, which loads through this same interface.
    """

    t: NDArray[np.float64]
    h: NDArray[np.float64]
    peak: float
    duration_hours: float
    scenario: str
    event_id: str
    native_dt: float
    provenance: dict = field(default_factory=dict)


def _si_time_axis(
    time_hours_arr: NDArray[np.float64],
) -> tuple[NDArray[np.float64], float]:
    """Validate a native-hours time axis and convert it to SI seconds.

    The shared axis seam of :func:`build_hydrograph_record` and the
    conditioning-record scaler: requires a 1-D, >= 2-sample, strictly
    increasing, uniformly spaced axis, and returns ``(t_seconds,
    native_dt_seconds)`` with ``native_dt`` derived from the actual converted
    spacing (never assumed 3600 s; ADR-0019 §6).

    Raises
    ------
    ValueError
        If the axis is not 1-D, has fewer than two samples (``native_dt``
        undefined), is not strictly increasing, or is not uniformly spaced.
    """
    if time_hours_arr.ndim != 1:
        raise ValueError("time_hours and discharge_m3s must be 1-D arrays.")
    if time_hours_arr.size < 2:
        raise ValueError(
            "need at least two samples to derive native_dt "
            f"(got {time_hours_arr.size})."
        )
    diffs_hours = np.diff(time_hours_arr)
    if np.any(diffs_hours <= 0.0):
        raise ValueError("time axis must be strictly increasing.")
    dt_hours = float(diffs_hours[0])
    if not np.allclose(diffs_hours, dt_hours, rtol=_SPACING_RTOL, atol=0.0):
        raise ValueError(
            "time axis must be uniformly spaced (native_dt is a single scalar "
            "resolution); got non-constant spacing."
        )
    # Unit boundary: hours -> SI seconds. native_dt is derived from the converted
    # spacing, so a non-hourly source yields the correct number of seconds.
    return time_hours_arr * _SECONDS_PER_HOUR, dt_hours * _SECONDS_PER_HOUR


def apply_rating_curve(
    discharge_m3s: ArrayLike, a_kp: float, b_kp: float
) -> NDArray[np.float64]:
    """Convert discharge Q(t) to river stage h(t) via Uemura thesis Eq. 4.19.

    ::

        h = sqrt(Q / a_kp) - b_kp

    Stage is in **m MSL** (ADR-0019 §3); ``-b_kp`` is an additive datum term,
    which is why b varies smoothly with bed elevation along the reach. This is
    **NOT** the power law ``a*Q**b`` — that literal reading of the email's
    one-line summary is wrong by ~120 orders of magnitude and must never be
    used (ADR-0019 §2).

    Pure and elementwise: a scalar Q maps to a scalar stage and a discharge
    array to a stage array of the same shape. This is the single conversion
    point; :func:`build_hydrograph_record` calls it, so any revision here
    propagates through the record with no separate code path.

    Parameters
    ----------
    discharge_m3s : array_like
        Discharge Q [m^3/s], as supplied by d4PDF (no unit scaling on Q);
        scalar or array. Must be non-negative.
    a_kp, b_kp : float
        Per-KP rating coefficients from non-uniform flow computation
        (``HQrelation_*Riv_2017.csv``); ``a_kp`` must be positive.

    Returns
    -------
    numpy.ndarray
        River stage h [m MSL], same shape as ``discharge_m3s``.

    Raises
    ------
    ValueError
        If ``a_kp`` is not positive or any ``Q / a_kp`` is negative — the
        ADR-0019 §2 radicand guard. It can never trip for physical inputs
        (Q >= 0, a_kp > 0); it exists to catch bad data loudly. Pinned as
        ``ValueError`` rather than a bare ``assert`` (``python -O`` strips
        asserts).
    """
    if not a_kp > 0.0:
        raise ValueError(f"a_kp must be positive (got {a_kp!r}).")
    discharge = np.asarray(discharge_m3s, dtype=np.float64)
    radicand = discharge / a_kp
    if np.any(radicand < 0.0):
        raise ValueError(
            "discharge must be non-negative: Q / a_kp must be non-negative "
            "before the square root (Eq. 4.19; ADR-0019 §2)."
        )
    return np.sqrt(radicand) - b_kp


def build_hydrograph_record(
    time_hours: ArrayLike,
    discharge_m3s: ArrayLike,
    *,
    a_kp: float,
    b_kp: float,
    scenario: str,
    event_id: str,
    provenance: Mapping | None = None,
) -> HydrographRecord:
    """Build a :class:`HydrographRecord` from a native-unit discharge series.

    The pure record-construction seam **and the unit boundary** (spec §1).
    Given time in the source's native unit (hours) and discharge in m^3/s, it:

    1. validates that the time axis is strictly increasing and uniformly
       spaced;
    2. converts the time axis to SI seconds (``t = time_hours * 3600``); the
       axis is NOT renumbered, so the d4PDF Time = 1..192 h window starts at
       t = 3600 s;
    3. derives ``native_dt`` [s] from the converted spacing (never assumed
       3600 s);
    4. applies :func:`apply_rating_curve` (Eq. 4.19) to produce h(t), uncapped
       (ADR-0019 §8); and
    5. records the peak stage, the elapsed duration in hours, and the
       provenance mapping.

    Deliberately independent of any file: it takes in-memory arrays, so the
    conversion and construction logic is testable without the large Excel
    workbooks (ADR-0019 Consequences; ``tests/test_hydrographs.py``).

    Parameters
    ----------
    time_hours : array_like, shape (T,)
        Time axis in **hours** (native source unit), strictly increasing and
        uniformly spaced; ``T >= 2`` so the spacing is defined.
    discharge_m3s : array_like, shape (T,)
        Discharge Q(t) [m^3/s], same length as ``time_hours``.
    a_kp, b_kp : float
        Per-KP Eq. 4.19 rating coefficients passed to
        :func:`apply_rating_curve`.
    scenario : str
        Climate scenario tag (``'historical'`` or ``'+4K'``); carried through.
    event_id : str
        Event identifier; carried through.
    provenance : mapping, optional
        ADR-0019 provenance (SST pattern, member ID, year, KP,
        ``discharge_proxied_from`` where applicable). Copied onto the record;
        omitted => ``{}`` so non-ensemble events (e.g. the 2016 observed
        hydrograph) use the same interface.

    Returns
    -------
    HydrographRecord
        The event with ``t``/``native_dt`` in seconds and ``h`` [m MSL] from
        Eq. 4.19.

    Raises
    ------
    ValueError
        If the time and discharge lengths differ, if fewer than two samples
        are given, if the time axis is not strictly increasing or not
        uniformly spaced (``native_dt`` would be ill-defined), or if the
        discharge fails the Eq. 4.19 radicand guard.
    """
    time_hours_arr = np.asarray(time_hours, dtype=np.float64)
    discharge = np.asarray(discharge_m3s, dtype=np.float64)

    if discharge.ndim != 1:
        raise ValueError("time_hours and discharge_m3s must be 1-D arrays.")
    t_seconds, native_dt = _si_time_axis(time_hours_arr)
    if time_hours_arr.shape != discharge.shape:
        raise ValueError(
            "time_hours and discharge_m3s must have the same length "
            f"(got {time_hours_arr.shape} and {discharge.shape})."
        )

    h = apply_rating_curve(discharge, a_kp, b_kp)
    duration_hours = float(time_hours_arr[-1] - time_hours_arr[0])

    return HydrographRecord(
        t=t_seconds,
        h=h,
        peak=float(h.max()),
        duration_hours=duration_hours,
        scenario=scenario,
        event_id=event_id,
        native_dt=float(native_dt),
        provenance=dict(provenance) if provenance is not None else {},
    )


def parse_member_header(header: str) -> dict:
    """Parse one d4PDF member column header into its provenance (ADR-0019 §1).

    Two forms exist, and the 50-flat-member past structure versus six-SST
    future structure must stay traceable:

    * past experiment ``HPB_mXXX_YYYY`` (member ID, calendar year; 3,000
      columns) -> scenario ``'historical'``, no SST pattern;
    * future experiment ``HFB_{SST}_mXXX_YYYY`` with SST one of CC, GF, HA,
      MI, MP, MR (15 members each; 5,400 columns) -> scenario ``'+4K'``.

    Parameters
    ----------
    header : str
        One member column header, verbatim from the ``QT`` sheet.

    Returns
    -------
    dict
        Keys ``experiment`` (``'HPB'``/``'HFB'``), ``scenario``
        (``'historical'``/``'+4K'``; ADR-0019 §9), ``sst`` (SST pattern, or
        None for HPB), ``member_id`` (verbatim ``'mXXX'``), ``year`` (int).

    Raises
    ------
    ValueError
        If the header matches neither form — including the ``Time`` column,
        an HFB header without a known SST pattern, or a missing year.
    """
    hpb = _HPB_HEADER_RE.fullmatch(header)
    if hpb is not None:
        return {
            "experiment": "HPB",
            "scenario": _SCENARIO_HISTORICAL,
            "sst": None,
            "member_id": hpb.group(1),
            "year": int(hpb.group(2)),
        }
    hfb = _HFB_HEADER_RE.fullmatch(header)
    if hfb is not None:
        return {
            "experiment": "HFB",
            "scenario": _SCENARIO_PLUS_4K,
            "sst": hfb.group(1),
            "member_id": hfb.group(2),
            "year": int(hfb.group(3)),
        }
    raise ValueError(
        f"not a d4PDF member header: {header!r}; expected 'HPB_mXXX_YYYY' or "
        f"'HFB_{{SST}}_mXXX_YYYY' with SST in {_SST_PATTERNS}."
    )


def resolve_discharge_source_kp(kp: float) -> tuple[float, str | None]:
    """Resolve which KP's discharge series a study node consumes (ADR-0019 §7).

    Discharge hydrograph files stop at KP 61.8 while rating coefficients
    extend to KP 62.8, so the five upper-Tokachi nodes (KP 62.0, 62.2, 62.4,
    62.6, 62.8) proxy the KP 61.8 discharge series — no major tributary enters
    between KP 61.8 and 62.8, so Q varies only gradually — while the rating,
    which is intensely local, stays each node's OWN ``a_kp``/``b_kp``. The
    proxy is provider-confirmed and flagged explicitly in the thesis text.

    Parameters
    ----------
    kp : float
        The study node's KP.

    Returns
    -------
    tuple of (float, str or None)
        ``(source_kp, marker)``: the KP whose discharge band file to read, and
        the ``discharge_proxied_from`` metadata marker (e.g. ``'KP61.8'``), or
        ``(kp, None)`` when the node has its own discharge coverage.
    """
    if any(abs(kp - node) <= _KP_MATCH_ATOL for node in _PROXY_NODE_KPS):
        return _PROXY_SOURCE_KP, f"KP{_PROXY_SOURCE_KP}"
    return kp, None


def validate_datum_consistency(record: HydrographRecord, z_toe_m: float) -> None:
    """Refuse to pair an M3 stage record with an unresolved z_toe datum.

    M3 stage records are referenced to **mean sea level** by construction
    (ADR-0019 §3; ~33-40 m MSL on the study reach). The generated configs now
    carry the ADR-0021 landside-toe elevations in m MSL, but a hand-built or
    legacy config could still pair an MSL record with the retired ADR-0018
    ``z_toe = 0.0`` placeholder. M8 computes the driving head as
    ``r_e * (h - z_toe)``, so feeding an MSL record against the placeholder
    would *silently* produce ~35 m heads — physically nonsensical, but
    numerically runnable. This guard makes that datum incompatibility loud
    instead: it is called wherever a real M3 record enters the M8 fragility
    path, and it raises on the placeholder.

    The check is necessary, not sufficient: it cannot prove an arbitrary
    nonzero z_toe is on the MSL datum, only refuse the known placeholder.

    Parameters
    ----------
    record : HydrographRecord
        A stage record produced by this module, i.e. h(t) in m MSL
        (ADR-0019 §3). Consulted for the error message only.
    z_toe_m : float
        The exit-point polder elevation the caller intends to hand M8
        (``geometry['z_toe']``), which must be on the same MSL datum.

    Raises
    ------
    ValueError
        If ``z_toe_m`` equals the PROVISIONAL 0.0 placeholder — the
        real-hydrograph path must not run until the datum is resolved.
    """
    if z_toe_m == _PROVISIONAL_Z_TOE_M:
        raise ValueError(
            "datum mismatch: this HydrographRecord's stages are m MSL "
            f"(ADR-0019 §3; event {record.event_id!r}, peak "
            f"{record.peak:.2f} m MSL) but z_toe is the retired PROVISIONAL "
            "0.0 placeholder (ADR-0018). Running M8 like this would "
            f"silently apply ~{record.peak:.0f} m driving heads. Set "
            "geometry.z_toe to the section's true exit-point elevation in "
            "m MSL (the ADR-0021 landside-toe values in the generated "
            "configs) before feeding real hydrographs to the engine."
        )


@dataclass(frozen=True)
class CanonicalShape:
    """A canonical d4PDF event's normalized stage shape at one study node.

    The G1 conditioning-level scaling artifact (ADR-0020 Decision 1): built
    once per run by :func:`load_canonical_shape` and consumed per level by
    :func:`conditioning_record_for_level`. The normalization is **stage
    domain** (not discharge — the Eq. 4.19 square root compresses, and the
    head structure is what M7 consumes), under the node's own local rating.

    Attributes
    ----------
    source_record : HydrographRecord
        The canonical event's full, unscaled stage record at the node (m MSL,
        SI axis, member provenance incl. the ADR-0019 §7 proxy marker and the
        resolved band/rating file names). Retained for provenance and for the
        run-level datum guard.
    shape : numpy.ndarray, shape (T,)
        The normalized stage shape ``(h - h_base) / (h_peak - h_base)``:
        exactly 0.0 at the base-flow floor and exactly 1.0 at the peak.
    h_base_m : float
        The section's base-flow stage [m MSL] — the trough baseline of the
        scaling rule (ADR-0021 Downstream-use item 4: h_base, NOT z_toe).
        Equals ``min(source_record.h)``, i.e. Eq. 4.19 at the band's constant
        base-flow discharge (75.44 m^3/s on the Tokachi study band) under the
        node's local rating.
    """

    source_record: HydrographRecord
    shape: NDArray[np.float64]
    h_base_m: float


def normalize_stage_shape(
    stage_m: ArrayLike,
) -> tuple[NDArray[np.float64], float, float]:
    """Normalize a stage series to its [base-flow, peak] span (G1 rule).

    Pure array operation: ``shape = (h - min(h)) / (max(h) - min(h))``, so
    ``min(shape) == 0.0`` and ``max(shape) == 1.0`` **exactly** (the extreme
    elements normalize to 0/1 with no float residue), which is what pins the
    rescaled trough floor at h_base exactly and the rescaled peak at h_i to
    within one ulp (the ``peak`` field itself is set to h_i verbatim).

    Parameters
    ----------
    stage_m : array_like, shape (T,)
        Stage series h(t) [m MSL].

    Returns
    -------
    tuple of (numpy.ndarray, float, float)
        ``(shape, h_base_m, h_peak_m)``: the normalized shape and the span it
        was normalized over.

    Raises
    ------
    ValueError
        If the series is constant (``max == min``) — a degenerate event has
        no shape to scale.
    """
    h = np.asarray(stage_m, dtype=np.float64)
    h_base = float(h.min())
    h_peak = float(h.max())
    if not h_peak > h_base:
        raise ValueError(
            "degenerate stage series: max(h) must exceed min(h) to define a "
            f"shape (got constant {h_base!r})."
        )
    return (h - h_base) / (h_peak - h_base), h_base, h_peak


def flood_timescales(stage_m: ArrayLike, dt_seconds: float) -> dict[str, float]:
    """Characteristic timescales of a stage hydrograph's dominant peak [s].

    Descriptive shape analysis used by the ADR-0032 aquifer-response diagnostic
    (spec §11): the rising-limb time is the denominator of the τ_aq/T_flood
    ratio, and the plateau width is what the native-resolution (Check B)
    Nyquist test interrogates. All widths are **integer multiples of
    ``dt_seconds``** — the native cadence itself — because they are counted on
    the record's own grid; that quantization is exactly the resolution the
    diagnostic asks about.

    Parameters
    ----------
    stage_m : array_like, shape (T,)
        Stage series h(t) [m above datum], uniformly sampled at ``dt_seconds``.
    dt_seconds : float
        Sampling interval [s] (the record's ``native_dt``).

    Returns
    -------
    dict of str to float
        ``rising_limb_s`` (10%-of-amplitude to the peak, on the final rising
        limb — the primary T_rise), ``rise_10_90_s`` (10%→90% rise time, a
        flashiness measure), ``plateau_s`` (time within 10% of the peak),
        ``fwhm_s`` (time above half amplitude), ``peak_m`` and ``amplitude_m``.

    Raises
    ------
    ValueError
        If the series is constant (no peak to characterize) or ``dt_seconds``
        is not positive.

    Notes
    -----
    Widths are measured on the **stage** record (post-rating), which is what
    the downstream initiation/progression modules consume; the Eq. 4.19 rating
    compresses the discharge peak, so the stage plateau is if anything broader
    than the discharge plateau. The rising-limb onset is the last crossing of
    the 10%-amplitude level before the peak, so a compound multi-peak record is
    characterized by the final approach to its global maximum.
    """
    if not dt_seconds > 0.0:
        raise ValueError(f"dt_seconds must be positive, got {dt_seconds!r}.")
    h = np.asarray(stage_m, dtype=np.float64)
    base = float(h.min())
    peak = float(h.max())
    amplitude = peak - base
    if not amplitude > 0.0:
        raise ValueError(
            "constant stage series has no dominant peak to characterize "
            f"(min == max == {base!r})."
        )
    shape = (h - base) / amplitude
    k_peak = int(np.argmax(h))
    pre = shape[: k_peak + 1]

    def _last_at_or_below(level: float, default: int) -> int:
        idx = np.where(pre <= level)[0]
        return int(idx[-1]) if idx.size else default

    k10 = _last_at_or_below(0.10, 0)
    k90 = _last_at_or_below(0.90, k_peak)
    return {
        "rising_limb_s": float((k_peak - k10) * dt_seconds),
        "rise_10_90_s": float(max(0, k90 - k10) * dt_seconds),
        "plateau_s": float(np.count_nonzero(shape >= 0.90) * dt_seconds),
        "fwhm_s": float(np.count_nonzero(shape >= 0.50) * dt_seconds),
        "peak_m": peak,
        "amplitude_m": amplitude,
    }


def conditioning_record_for_level(
    canonical: CanonicalShape, level_m: float, *, scenario: str
) -> HydrographRecord:
    """Scale the canonical shape to conditioning level ``level_m`` (G1 rule).

    The pure per-level scaler behind the fragility sweep (ADR-0020 Decision 1)::

        h(t) = h_base + (level_m - h_base) * shape(t)

    so the trough floor stays pinned at the section's base-flow stage h_base
    (ADR-0021 Downstream-use item 4 — NOT at z_toe, which would artificially
    deepen every inter-peak recession and bias the memory-model arrest) while
    the peak scales to ``level_m``. No time rescaling: the full source window
    at its native resolution (hourly, ADR-0019 §6). ``peak`` is set to
    ``level_m`` **verbatim** — the authoritative conditioning anchor (ADR-0010;
    M8 ambiguity 3); ``max(h)`` equals it to within one ulp.

    A level at or below h_base has no positive event amplitude to scale;
    such (sub-base-flow) conditioning levels are emitted as a **constant
    stage** ``h(t) = level_m`` — the zero-load floor of the fragility curve
    (on the study sections h_base already sits 1.7-3.5 m below the landside
    toe, so these levels produce no blanket overpressure either way).

    Deterministic pure function of its arguments (no RNG, no I/O): with the
    canonical shape loaded once in the main process, the parallel sweep stays
    bit-identical to a serial run (the run.py reproducibility guarantee).

    Parameters
    ----------
    canonical : CanonicalShape
        The event shape from :func:`load_canonical_shape`.
    level_m : float
        Conditioning level h_i [m MSL]; becomes ``peak`` verbatim.
    scenario : str
        The **run's** climate scenario tag, carried onto the record (the
        shape source event's own scenario stays in provenance — the same
        canonical HPB shapes drive both scenarios by design; climate enters
        via the loading distribution downstream, not the fragility shape).

    Returns
    -------
    HydrographRecord
        The level's loading record: source time axis and resolution,
        rescaled stages, ``peak == level_m``, and provenance carrying the
        member info plus ``shape_source_event``, ``conditioning_level_m_msl``
        and ``h_base_m_msl``.
    """
    source = canonical.source_record
    level = float(level_m)
    if level <= canonical.h_base_m:
        h = np.full_like(source.h, level)
    else:
        h = canonical.h_base_m + (level - canonical.h_base_m) * canonical.shape
    provenance = {
        **source.provenance,
        "shape_source_event": source.event_id,
        "conditioning_level_m_msl": level,
        "h_base_m_msl": canonical.h_base_m,
    }
    return HydrographRecord(
        t=source.t,
        h=h,
        peak=level,
        duration_hours=source.duration_hours,
        scenario=scenario,
        event_id=f"{source.event_id}_scaled_h{level:g}",
        native_dt=source.native_dt,
        provenance=provenance,
    )


def resample_record(
    record: HydrographRecord, target_dt_seconds: float
) -> HydrographRecord:
    """Refine a record onto a finer uniform time grid (ADR-0013 resample hook).

    Linear interpolation of the stage series onto ``target_dt_seconds``, the
    record-construction resampling that ADR-0013 assigns to M3 and ADR-0022
    anticipated (its decision 3 forward requirement; the Phase 2 native/2
    replay and the ADR-0030 Phase 1 integration-Δt policy both consume it).
    Interpolation adds **no information** to the hourly d4PDF source — it only
    refines the forward-Euler integration grid, which matters because at
    3600 s a single Euler step can jump the H_eq equilibrium barrier for
    high-C_e·k_aq realizations (ADR-0030): the loading signal stays the
    resolved hourly signal.

    The refinement is restricted to **integer subdivisions** of the native
    grid so every native sample is a node of the new grid: the resampled
    series passes through all source points exactly, ``max(h)`` (hence
    ``peak``) is preserved, and the ``native_dt``-halving ladder of the §11
    convergence test is exactly nested.

    Parameters
    ----------
    record : HydrographRecord
        The source record (any uniform grid).
    target_dt_seconds : float
        The target resolution [s]; must satisfy ``record.native_dt =
        k * target_dt_seconds`` for a positive integer k. ``k == 1`` returns
        the record unchanged.

    Returns
    -------
    HydrographRecord
        The refined record: ``native_dt = target_dt_seconds``, same span,
        ``peak``/``duration_hours``/``scenario``/``event_id`` unchanged, and
        provenance extended with ``resampled_from_native_dt_s`` and
        ``resample_factor``.

    Raises
    ------
    ValueError
        If ``target_dt_seconds`` is not positive, exceeds ``native_dt``, or
        does not divide it to an integer refinement factor.
    """
    target = float(target_dt_seconds)
    if target <= 0.0:
        raise ValueError(f"target_dt_seconds must be > 0 (got {target}).")
    factor_float = record.native_dt / target
    factor = int(round(factor_float))
    if factor < 1 or not math.isclose(factor_float, factor, rel_tol=1e-9):
        raise ValueError(
            f"target_dt_seconds={target} must be an integer subdivision of "
            f"native_dt={record.native_dt} (native_dt / target = "
            f"{factor_float:.6g}); coarsening or non-nested grids would move "
            "the loading signal, not just the integration grid (ADR-0013)."
        )
    if factor == 1:
        return record
    n_refined = (record.t.size - 1) * factor + 1
    t_refined = record.t[0] + np.arange(n_refined, dtype=np.float64) * target
    h_refined = np.interp(t_refined, record.t, record.h)
    return HydrographRecord(
        t=t_refined,
        h=h_refined,
        peak=record.peak,
        duration_hours=record.duration_hours,
        scenario=record.scenario,
        event_id=record.event_id,
        native_dt=target,
        provenance={
            **record.provenance,
            "resampled_from_native_dt_s": record.native_dt,
            "resample_factor": factor,
        },
    )


def experiment_for_scenario(scenario: str) -> str:
    """Map a climate scenario tag to its d4PDF experiment (ADR-0020 §3).

    ``'historical'`` -> ``'HPB'`` (past experiment), ``'+4K'`` -> ``'HFB'``
    (future experiment; ADR-0019 §9). This function is the single source of
    that mapping: the orchestrator selects band workbooks from
    ``config.scenario`` through here, and no config field duplicates it.

    Parameters
    ----------
    scenario : str
        Climate scenario tag.

    Returns
    -------
    str
        ``'HPB'`` or ``'HFB'``.

    Raises
    ------
    ValueError
        If the scenario is neither ``'historical'`` nor ``'+4K'``.
    """
    try:
        return _EXPERIMENT_BY_SCENARIO[scenario]
    except KeyError:
        raise ValueError(
            f"unknown scenario {scenario!r}; expected one of "
            f"{sorted(_EXPERIMENT_BY_SCENARIO)} (ADR-0019 §9)."
        ) from None


def rating_curve_path(data_root: str | Path, river: str) -> Path:
    """Return the rating-coefficient CSV path for a river (ADR-0020 §2).

    One naming convention, one place:
    ``{data_root}/rating_curves/HQrelation_{river}Riv_2017.csv``. The path is
    derived, not configured per file; existence is checked by the reader
    (:func:`load_rating_coefficients`) when the file is opened.

    Parameters
    ----------
    data_root : str or pathlib.Path
        Root of the raw data drop.
    river : str
        ``'Tokachi'`` or ``'Satsunai'``.

    Returns
    -------
    pathlib.Path
        The per-river rating CSV path.

    Raises
    ------
    ValueError
        If ``river`` is not one of the known rivers (the name is interpolated
        into a filename, so it is validated rather than trusted).
    """
    if river not in _KNOWN_RIVERS:
        raise ValueError(f"unknown river {river!r}; expected one of {_KNOWN_RIVERS}.")
    return Path(data_root) / "rating_curves" / f"HQrelation_{river}Riv_2017.csv"


def resolve_band_workbook(
    data_root: str | Path, *, river: str, kp: float, scenario: str
) -> Path:
    """Resolve the unique band workbook for (river, kp, scenario) (ADR-0020 §4).

    Resolution rule, in order:

    1. :func:`resolve_discharge_source_kp` first, so the upper-Tokachi nodes
       KP 62.0-62.8 route to the KP 61.8 band (ADR-0019 §7) while their
       rating stays their own (handled separately by
       :func:`load_hydrograph_ensemble`).
    2. Scan ``{data_root}/hydrographs/`` (sorted, so the scan is
       filesystem-order independent) for filenames matching the provider's
       band grammar, parsed — not hardcoded — from the name.
    3. Keep the files whose river and experiment
       (:func:`experiment_for_scenario`) match and whose inclusive
       ``[lo, hi]`` KP range covers the (proxied) source KP; exactly one
       must remain.

    Pure function of its arguments plus the directory listing: no RNG, no
    fallback, so the ``run.py`` reproducibility-by-construction guarantee is
    preserved (ADR-0020 §5).

    Parameters
    ----------
    data_root : str or pathlib.Path
        Root of the raw data drop (expects a ``hydrographs/`` subdirectory).
    river : str
        ``'Tokachi'`` or ``'Satsunai'``.
    kp : float
        The study node's KP (pre-proxy; the §7 routing is applied here).
    scenario : str
        Climate scenario tag; selects the experiment (HPB/HFB).

    Returns
    -------
    pathlib.Path
        The single matching band workbook.

    Raises
    ------
    ValueError
        If the ``hydrographs/`` directory is missing, if no band file covers
        the KP for this river/experiment (no silent nearest-band fallback),
        or if more than one does (ambiguous drop — overlapping bands must be
        resolved by the provider, not guessed at).
    """
    experiment = experiment_for_scenario(scenario)
    if river not in _KNOWN_RIVERS:
        raise ValueError(f"unknown river {river!r}; expected one of {_KNOWN_RIVERS}.")
    source_kp, _ = resolve_discharge_source_kp(kp)

    hydro_dir = Path(data_root) / "hydrographs"
    if not hydro_dir.is_dir():
        raise ValueError(
            f"no hydrographs/ directory under {data_root!s}; expected the d4PDF "
            "band workbooks there (ADR-0020 §1)."
        )

    matches: list[Path] = []
    for path in sorted(hydro_dir.iterdir()):
        parsed = _BAND_WORKBOOK_RE.fullmatch(path.name)
        if parsed is None:
            continue
        if parsed["experiment"] != experiment or parsed["river"] != river:
            continue
        lo, hi = float(parsed["lo"]), float(parsed["hi"])
        if lo - _KP_MATCH_ATOL <= source_kp <= hi + _KP_MATCH_ATOL:
            matches.append(path)

    if not matches:
        raise ValueError(
            f"no {experiment} band workbook covers KP {source_kp:g} for the "
            f"{river} in {hydro_dir!s} (node KP {kp:g}; ADR-0020 §4 — no "
            "nearest-band fallback)."
        )
    if len(matches) > 1:
        raise ValueError(
            f"ambiguous band coverage for {river} KP {source_kp:g} "
            f"({experiment}): {[p.name for p in matches]!r} all match; "
            "overlapping bands in the data drop must be resolved, not guessed."
        )
    return matches[0]


# ============================================================================
# FILE-READING SEAM — thin, dumb parsers. No unit conversion, no rating-curve
# application, no metadata derivation: those belong to the pure layer above,
# which is what keeps the physics-adjacent logic testable without the large
# Excel workbooks (ADR-0019 Consequences).
# ============================================================================
def load_rating_coefficients(path: str | Path) -> dict[float, tuple[float, float]]:
    """Read one ``HQrelation_*Riv_2017.csv`` into a KP -> (a_kp, b_kp) mapping.

    The rating-coefficient CSVs are **Shift-JIS encoded** (NOT UTF-8; the
    encoding is specified explicitly here) with header cells ``HQ_a``/``HQ_b``
    containing **full-width** characters, and columns River, KP, a, b at
    0.2 km spacing (ADR-0019 §5). Header cells are NFKC-normalized before
    matching, so the coefficient columns are bound by *name*, not position —
    a reordered or re-exported file cannot silently swap a and b.

    Parameters
    ----------
    path : str or pathlib.Path
        One per-river rating CSV (Tokachi or Satsunai).

    Returns
    -------
    dict of float -> tuple of (float, float)
        ``{kp: (a_kp, b_kp)}`` in the file's units (a in m^3/s scaling, b in
        m MSL datum offset), covering the full study reach and beyond.

    Raises
    ------
    ValueError
        If the header lacks the KP / HQ_a / HQ_b columns.
    """
    with open(path, encoding="shift_jis", newline="") as stream:
        rows = list(csv.reader(stream))
    if not rows:
        raise ValueError(f"empty rating-coefficient CSV: {path}")

    header = [unicodedata.normalize("NFKC", cell).strip() for cell in rows[0]]
    try:
        kp_idx = header.index(_RATING_KP_HEADER)
        a_idx = header.index(_RATING_A_HEADER)
        b_idx = header.index(_RATING_B_HEADER)
    except ValueError as exc:
        raise ValueError(
            f"rating CSV header must contain {_RATING_KP_HEADER!r}, "
            f"{_RATING_A_HEADER!r} and {_RATING_B_HEADER!r} after NFKC "
            f"normalization; got {header!r} in {path}."
        ) from exc

    coefficients: dict[float, tuple[float, float]] = {}
    for row in rows[1:]:
        if not row or all(not cell.strip() for cell in row):
            continue
        coefficients[float(row[kp_idx])] = (float(row[a_idx]), float(row[b_idx]))
    return coefficients


def read_discharge_ensemble(
    path: str | Path,
) -> tuple[NDArray[np.float64], dict[str, NDArray[np.float64]]]:
    """Read one d4PDF discharge workbook's ``QT`` sheet (thin parser).

    One workbook covers one river/KP-band; a single discharge column applies
    across the whole band (ADR-0019 §1). Layout: sheet ``QT``, column 1
    ``Time`` (integer hours, 1..192), each subsequent column one ensemble
    member's Q(t) in m^3/s under an ``HPB_``/``HFB_`` header.

    Opened with openpyxl in **read-only mode**: the workbooks carry 3,000
    (HPB) or 5,400 (HFB) member columns, and read-only streaming avoids
    materializing the full worksheet object tree.

    Dumb parser by design: returns raw columns in **native source units**
    (time in hours, Q in m^3/s) with member headers verbatim, and does no unit
    conversion, header parsing, or rating-curve application — those belong to
    the pure layer (:func:`build_hydrograph_record`,
    :func:`parse_member_header`).

    Parameters
    ----------
    path : str or pathlib.Path
        One per-river/KP-band ``.xlsx`` discharge workbook.

    Returns
    -------
    tuple of (numpy.ndarray, dict of str -> numpy.ndarray)
        ``(time_hours, members)``: the shared time axis [hours] and one
        discharge series [m^3/s] per verbatim member header, all of equal
        length.

    Raises
    ------
    ValueError
        If the ``QT`` sheet is missing, the first column is not ``Time``, no
        member columns exist, or a discharge cell is missing/non-numeric.
    """
    # Imported here so the pure conversion layer stays importable without the
    # Excel dependency.
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if _QT_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"workbook {path} has no {_QT_SHEET_NAME!r} sheet "
                f"(found {workbook.sheetnames!r})."
            )
        rows = workbook[_QT_SHEET_NAME].iter_rows(values_only=True)

        header_row = next(rows, None)
        if header_row is None:
            raise ValueError(f"sheet {_QT_SHEET_NAME!r} in {path} is empty.")
        first_header = str(header_row[0]).strip() if header_row[0] is not None else ""
        if first_header != _TIME_COLUMN_HEADER:
            raise ValueError(
                f"first column of sheet {_QT_SHEET_NAME!r} must be "
                f"{_TIME_COLUMN_HEADER!r} (got {first_header!r}) in {path}."
            )
        member_columns = [
            (idx, str(cell).strip())
            for idx, cell in enumerate(header_row[1:], start=1)
            if cell is not None and str(cell).strip()
        ]
        if not member_columns:
            raise ValueError(
                f"sheet {_QT_SHEET_NAME!r} in {path} has no member columns."
            )

        time_values: list[float] = []
        q_rows: list[list[float]] = []
        for row in rows:
            if not row or row[0] is None:
                continue  # trailing padding rows in read-only streaming
            time_values.append(float(row[0]))
            q_row: list[float] = []
            for idx, name in member_columns:
                value = row[idx] if idx < len(row) else None
                if value is None:
                    raise ValueError(
                        f"missing discharge for member {name!r} at "
                        f"{_TIME_COLUMN_HEADER}={row[0]!r} in {path}."
                    )
                q_row.append(float(value))
            q_rows.append(q_row)
    finally:
        workbook.close()

    time_hours = np.asarray(time_values, dtype=np.float64)
    q_matrix = np.asarray(q_rows, dtype=np.float64)
    members = {
        name: q_matrix[:, column] for column, (_, name) in enumerate(member_columns)
    }
    return time_hours, members


def load_hydrograph_ensemble(
    path: str | Path,
    *,
    kp: float,
    rating_coefficients: Mapping[float, tuple[float, float]],
) -> dict[str, HydrographRecord]:
    """Load one discharge workbook as stage records for one study node.

    The spec §2 composition ``dict[event_id -> HydrographRecord]``: read the
    band workbook once (:func:`read_discharge_ensemble`), then per member
    parse the header provenance and build the stage record through the single
    Eq. 4.19 conversion path. The rating coefficients are looked up at the
    node's **own** KP from the supplied mapping — never at the discharge
    band's KP — so the ADR-0019 §7 rule (proxied discharge, local rating) is
    enforced structurally rather than by caller discipline.

    Upper-Tokachi proxy (ADR-0019 §7): for KP 62.0-62.8 the caller passes the
    **KP 61.8 band workbook** as ``path`` (use
    :func:`resolve_discharge_source_kp` to pick it); this function then stamps
    ``discharge_proxied_from='KP61.8'`` into each record's provenance
    automatically while applying the node's own local coefficients.

    Parameters
    ----------
    path : str or pathlib.Path
        The ``.xlsx`` discharge workbook whose band covers this node's
        discharge source KP.
    kp : float
        The study node's KP; selects the rating coefficients and is recorded
        in provenance.
    rating_coefficients : mapping of float -> tuple of (float, float)
        Per-KP ``(a_kp, b_kp)``, e.g. from :func:`load_rating_coefficients`.

    Returns
    -------
    dict of str -> HydrographRecord
        One record per ensemble member, keyed by the verbatim column header
        (also the record's ``event_id``). Each record's provenance carries
        experiment, scenario, SST pattern (None for HPB), member ID, year,
        ``kp``, and ``discharge_proxied_from`` where applicable.

    Raises
    ------
    ValueError
        If no rating coefficients exist for ``kp``, or on any reader/parser
        validation failure.
    """
    try:
        a_kp, b_kp = rating_coefficients[kp]
    except KeyError:
        raise ValueError(
            f"no rating coefficients for KP {kp}; the rating files cover the "
            "full study reach (ADR-0019 §5), so a miss means the wrong river's "
            "file or a malformed KP."
        ) from None

    _, discharge_proxied_from = resolve_discharge_source_kp(kp)

    time_hours, members = read_discharge_ensemble(path)
    records: dict[str, HydrographRecord] = {}
    for header, discharge_m3s in members.items():
        info = parse_member_header(header)
        provenance = {**info, "kp": kp}
        if discharge_proxied_from is not None:
            provenance["discharge_proxied_from"] = discharge_proxied_from
        records[header] = build_hydrograph_record(
            time_hours,
            discharge_m3s,
            a_kp=a_kp,
            b_kp=b_kp,
            scenario=str(info["scenario"]),
            event_id=header,
            provenance=provenance,
        )
    return records


def load_canonical_shape(
    data_root: str | Path, *, river: str, kp: float, event_id: str
) -> CanonicalShape:
    """Load one canonical event's normalized shape at one study node (G1).

    The once-per-run composition behind the conditioning-level sweep
    (ADR-0020): resolve the rating CSV and the band workbook, read the single
    canonical member's discharge, build its stage record at the node's OWN
    rating (the ADR-0019 §7 proxy rule holds: an upper-Tokachi node reads the
    KP 61.8 band but keeps its local coefficients, and the record is stamped
    ``discharge_proxied_from``), and normalize it in stage domain
    (:func:`normalize_stage_shape`).

    The band workbook is selected by the **event's own experiment** (parsed
    from ``event_id``), not the run's scenario: the approved canonical shapes
    are HPB members and drive both the historical and the +4K fragility runs
    (climate enters via the loading distribution downstream, not the shape).

    The returned shape is a pure value: everything downstream
    (:func:`conditioning_record_for_level`) is deterministic, so loading once
    in the main process preserves the run.py parallel == serial guarantee
    (ADR-0020 §5).

    Parameters
    ----------
    data_root : str or pathlib.Path
        Root of the raw data drop (ADR-0020 layout).
    river : str
        ``'Tokachi'`` or ``'Satsunai'``.
    kp : float
        The study node's KP; selects the local rating and (after §7 proxy
        routing) the band workbook.
    event_id : str
        The canonical member's verbatim column header (ADR-0019 §1 grammar),
        e.g. ``config.hydrograph_source.canonical_event_ids[0]``.

    Returns
    -------
    CanonicalShape
        The normalized shape, its base-flow stage h_base [m MSL], and the
        full source record (provenance includes the member info, the node
        KP, the proxy marker where applicable, and the resolved
        ``band_workbook`` / ``rating_csv`` file names).

    Raises
    ------
    ValueError
        If the event ID is not a member header, the rating lacks the KP, the
        band resolution fails, or the workbook does not contain the event.
    """
    info = parse_member_header(event_id)  # validates + gives the experiment
    rating_csv = rating_curve_path(data_root, river)
    rating_coefficients = load_rating_coefficients(rating_csv)
    try:
        a_kp, b_kp = rating_coefficients[kp]
    except KeyError:
        raise ValueError(
            f"no rating coefficients for KP {kp} in {rating_csv.name}; the "
            "rating files cover the full study reach (ADR-0019 §5), so a miss "
            "means the wrong river or a malformed KP."
        ) from None

    workbook = resolve_band_workbook(
        data_root, river=river, kp=kp, scenario=str(info["scenario"])
    )
    _, discharge_proxied_from = resolve_discharge_source_kp(kp)

    time_hours, members = read_discharge_ensemble(workbook)
    try:
        discharge_m3s = members[event_id]
    except KeyError:
        raise ValueError(
            f"canonical event {event_id!r} not found in {workbook.name} "
            f"({len(members)} member columns)."
        ) from None

    provenance = {
        **info,
        "kp": kp,
        "band_workbook": workbook.name,
        "rating_csv": rating_csv.name,
    }
    if discharge_proxied_from is not None:
        provenance["discharge_proxied_from"] = discharge_proxied_from

    source_record = build_hydrograph_record(
        time_hours,
        discharge_m3s,
        a_kp=a_kp,
        b_kp=b_kp,
        scenario=str(info["scenario"]),
        event_id=event_id,
        provenance=provenance,
    )
    shape, h_base_m, _ = normalize_stage_shape(source_record.h)
    return CanonicalShape(source_record=source_record, shape=shape, h_base_m=h_base_m)
