"""Tests for M3 ``hydrograph_loader`` (``hydrographs.py``) per ADR-0019.

These tests supersede the interface-first ``tests/test_m3.py`` suite, which was
written **before** the d4PDF data format and the H-Q conversion were confirmed
and assumed a provisional power-law rating curve. The authoritative record is
now ``docs/decisions/0019-m3-hydrograph-data-and-hq-conversion.md`` (Uemura
thesis Eq. 4.19 + July 2026 correspondence); every golden value here traces to
that ADR or to the in-repo rating-coefficient CSVs it describes.

Authoritative facts pinned by this suite (ADR-0019 section in parentheses)
---------------------------------------------------------------------------
* H-Q conversion is Uemura thesis Eq. 4.19, ``h_t = sqrt(Q_t / a_kp) - b_kp``,
  with per-KP coefficients; it is **NOT** a power law ``a*Q**b`` (§2).
* ``-b_kp`` is an additive datum term; stage is in **m MSL**, the same datum as
  the 2019 bank-height HWL values (§2, §3).
* Obihiro anchor (Tokachi KP 56.6): a = 140.33, b = -32.49, HWL = 38.14 m MSL;
  Q = ~4,180 m^3/s reproduces the 2016 peak of ~37.95 m MSL, 0.19 m below HWL,
  and Q = 3,000-7,500 m^3/s spans 37.1-39.8 m MSL (§4).
* Rating-coefficient CSVs are Shift-JIS with full-width ``HQ_a``/``HQ_b``
  header cells and columns River, KP, a, b; coverage includes the upper
  Tokachi nodes KP 62.0-62.8 (§5, §7).
* Native resolution is 1 hour; ``native_dt`` = 3600 s is **derived** from the
  Time column (integer hours 1..192), and the time axis is converted to SI
  seconds inside M3 (§1, §6).
* Member column headers ``HPB_mXXX_YYYY`` / ``HFB_{SST}_mXXX_YYYY`` are parsed
  and their provenance (SST pattern, member ID, year) preserved; HPB maps to
  scenario ``'historical'`` and HFB to ``'+4K'`` (§1, §9).
* Stage is **not** capped at the levee crest inside M3: capping belongs to
  Uemura's overflow mechanism, not the seepage boundary condition (§8).

Pinned M3 public interface
--------------------------
``HydrographRecord``
    Frozen record with the spec §2 fields ``t`` (s), ``h`` (m MSL), ``peak``,
    ``duration_hours``, ``scenario``, ``event_id``, ``native_dt``, plus the
    ADR-0019 ``provenance`` dict (SST pattern where applicable, member ID,
    year, KP, ``discharge_proxied_from`` where applicable). ``provenance``
    defaults to ``{}`` so the 2016 observed hydrograph loads through the same
    interface (ADR-0019 Consequences).

``apply_rating_curve(discharge_m3s, a_kp, b_kp) -> ndarray``
    Pure, elementwise Eq. 4.19 conversion. Raises ``ValueError`` if the
    radicand ``Q/a_kp`` is negative (the ADR-0019 §2 guard).

``build_hydrograph_record(time_hours, discharge_m3s, *, a_kp, b_kp, scenario,
    event_id, provenance=None) -> HydrographRecord``
    The pure record-construction seam and the unit boundary: converts the time
    axis from source hours to SI seconds, derives ``native_dt`` from the
    actual (converted) spacing, applies the rating curve, and validates that
    the axis is strictly increasing and uniformly spaced.

``load_rating_coefficients(path) -> dict[float, tuple[float, float]]``
    Reads one ``HQrelation_*Riv_2017.csv`` (Shift-JIS, full-width header
    cells) into a KP -> (a_kp, b_kp) mapping.

``parse_member_header(header) -> dict``
    Parses one ensemble-member column header into provenance with keys
    ``experiment``, ``scenario``, ``sst`` (None for HPB), ``member_id``,
    ``year``; rejects non-member headers (e.g. ``Time``) with ``ValueError``.

File-reading seam (ADR-0019 Consequences)
-----------------------------------------
The large per-band Excel files (sheet ``QT``) are deliberately **not** read by
this suite: the Excel reader stays a thin seam whose pure parts — header
parsing, H-Q conversion, record construction — are locked here without any
large file on disk. The small rating-coefficient CSVs *are* read: hermetically
via a byte-exact Shift-JIS fixture, and directly from ``data/raw`` where those
(untracked) files are present, which is where the Obihiro anchor is asserted
end-to-end.
"""

from __future__ import annotations

import dataclasses
from pathlib import Path

import numpy as np
import pytest

from bep_reliability_engine.hydrographs import (
    CanonicalShape,
    HydrographRecord,
    apply_rating_curve,
    build_hydrograph_record,
    conditioning_record_for_level,
    experiment_for_scenario,
    load_canonical_shape,
    load_hydrograph_ensemble,
    load_rating_coefficients,
    normalize_stage_shape,
    parse_member_header,
    rating_curve_path,
    read_discharge_ensemble,
    resample_record,
    resolve_band_workbook,
    validate_datum_consistency,
)

_SECONDS_PER_HOUR = 3600.0

# Fixture rating coefficients (Eq. 4.19): with a_kp = 4.0 the fixture's
# discharge values make Q/a_kp a perfect square, and b_kp = -30.0 adds a
# datum-like +30 m, so every expected stage is exact (no float fuzz).
_A_KP = 4.0
_B_KP = -30.0

# Obihiro anchor, ADR-0019 §4/§5: Tokachi KP 56.6 coefficients and HWL (m MSL).
_OBIHIRO_A_KP = 140.33
_OBIHIRO_B_KP = -32.49
_OBIHIRO_HWL_M_MSL = 38.14
_OBIHIRO_PEAK_Q_M3S = 4180.0

# Real rating-coefficient CSVs (ADR-0019 §5). data/raw is not git-tracked, so
# the tests that read the genuine files skip on clones without the data drop.
_RATING_DIR = Path(__file__).resolve().parents[1] / "data" / "raw" / "rating_curves"
_TOKACHI_CSV = _RATING_DIR / "HQrelation_TokachiRiv_2017.csv"
_SATSUNAI_CSV = _RATING_DIR / "HQrelation_SatsunaiRiv_2017.csv"

requires_rating_csvs = pytest.mark.skipif(
    not (_TOKACHI_CSV.exists() and _SATSUNAI_CSV.exists()),
    reason="rating-coefficient CSVs (untracked data/raw) not present",
)

# Real d4PDF discharge workbooks (ADR-0019 §1). Like the rating CSVs these live
# in untracked data/raw, so the file-seam smoke tests skip on fresh clones.
_HYDROGRAPH_DIR = Path(__file__).resolve().parents[1] / "data" / "raw" / "hydrographs"
_HPB_UPPER_TOKACHI_XLSX = (
    _HYDROGRAPH_DIR / "Hydro Data, HPB, Tokachi Riv. KP056.20-KP061.80.xlsx"
)
_HFB_UPPER_TOKACHI_XLSX = (
    _HYDROGRAPH_DIR / "Hydro Data, HFB, Tokachi Riv. KP056.20-KP061.80.xlsx"
)

requires_hpb_workbook = pytest.mark.skipif(
    not (_HPB_UPPER_TOKACHI_XLSX.exists() and _TOKACHI_CSV.exists()),
    reason="d4PDF HPB workbook / rating CSV (untracked data/raw) not present",
)
requires_hfb_workbook = pytest.mark.skipif(
    not _HFB_UPPER_TOKACHI_XLSX.exists(),
    reason="d4PDF HFB workbook (untracked data/raw) not present",
)

# The six prescribed sea-surface-temperature patterns of the +4K (HFB)
# experiment, 15 members each (ADR-0019 §1).
_SST_PATTERNS = ["CC", "GF", "HA", "MI", "MP", "MR"]


def _synthetic_hourly_discharge() -> tuple[np.ndarray, np.ndarray]:
    """Return an in-memory synthetic (time_hours, discharge_m3s) hydrograph.

    A single-peaked hourly discharge series in **native source units**: 25
    samples at 1-hour spacing (24-hour span) with a clear interior maximum of
    1600 m^3/s. Discharges are ``_A_KP`` times perfect squares, so under
    Eq. 4.19 with (a_kp, b_kp) = (4.0, -30.0) every stage is exact:
    h = sqrt(Q/4) + 30, rising from 31.0 to the peak 50.0 m.

    Returns
    -------
    time_hours : numpy.ndarray, shape (25,)
        Time axis in **hours** (native source unit), ``np.arange(25)``.
    discharge_m3s : numpy.ndarray, shape (25,)
        Discharge Q(t) [m^3/s] with a single interior peak of 1600.
    """
    time_hours = np.arange(25, dtype=np.float64)
    squares = np.array(
        [1.0, 4.0, 9.0, 16.0, 25.0, 36.0, 49.0, 64.0, 81.0, 100.0, 225.0, 324.0],
        dtype=np.float64,
    )
    ramp = _A_KP * squares
    discharge = np.concatenate([ramp, [_A_KP * 400.0], ramp[::-1]])
    assert discharge.shape == time_hours.shape
    return time_hours, discharge


# --- HydrographRecord field contract -----------------------------------------
def test_hydrograph_record_field_contract() -> None:
    """The record exposes the spec §2 fields plus ADR-0019 ``provenance``.

    Locks the schema M8/run.py duck-type against: ``t``/``h`` equal-length
    float arrays, scalar ``peak``/``duration_hours``/``native_dt``, string
    ``scenario``/``event_id``, and the provenance dict of ADR-0019.
    """
    time_hours, discharge = _synthetic_hourly_discharge()
    record = build_hydrograph_record(
        time_hours,
        discharge,
        a_kp=_A_KP,
        b_kp=_B_KP,
        scenario="historical",
        event_id="synthetic_0001",
    )

    assert isinstance(record, HydrographRecord)
    expected_fields = {
        "t",
        "h",
        "peak",
        "duration_hours",
        "scenario",
        "event_id",
        "native_dt",
        "provenance",
    }
    assert {f.name for f in dataclasses.fields(record)} == expected_fields

    # Array fields: equal-length 1-D float arrays.
    assert isinstance(record.t, np.ndarray)
    assert isinstance(record.h, np.ndarray)
    assert record.t.shape == record.h.shape == discharge.shape
    assert np.issubdtype(record.t.dtype, np.floating)
    assert np.issubdtype(record.h.dtype, np.floating)

    # Scalar metadata fields (np.float64 is a subclass of float).
    assert isinstance(record.peak, float)
    assert isinstance(record.duration_hours, float)
    assert isinstance(record.native_dt, float)

    # Tag and provenance fields.
    assert isinstance(record.scenario, str)
    assert isinstance(record.event_id, str)
    assert record.event_id == "synthetic_0001"
    assert isinstance(record.provenance, dict)


# --- Q -> h conversion: Uemura thesis Eq. 4.19 (ADR-0019 §2) ------------------
@pytest.mark.parametrize(
    ("a_kp", "b_kp", "discharge", "expected_stage"),
    [
        # sqrt(100/4) - (-10) = 5 + 10 = 15. A power-law reading a*Q**b would
        # give 4 * 100**-10 = 4e-20 — the two forms cannot be confused.
        (4.0, -10.0, 100.0, 15.0),
        # b_kp = 0: pure sqrt(Q/a). sqrt(50/2) = 5.
        (2.0, 0.0, 50.0, 5.0),
        # Positive b_kp SUBTRACTS (sign-convention guard): sqrt(16/1) - 5 = -1.
        (1.0, 5.0, 16.0, -1.0),
        # Obihiro anchor: sqrt(4180/140.33) + 32.49 = 5.457740... + 32.49.
        (140.33, -32.49, 4180.0, 37.947740),
    ],
)
def test_rating_curve_eq_4_19_hand_values(
    a_kp: float, b_kp: float, discharge: float, expected_stage: float
) -> None:
    """``apply_rating_curve`` reproduces hand-computed h = sqrt(Q/a_kp) - b_kp.

    This is the M3 formula guard (ADR-0019 §2): the conversion is Uemura
    thesis Eq. 4.19 exactly, NOT the power law h = a*Q**b of the email's
    one-line summary, which is wrong by ~120 orders of magnitude. Each
    expected value is computable with a calculator.
    """
    result = apply_rating_curve(discharge, a_kp, b_kp)
    assert float(result) == pytest.approx(expected_stage, abs=1e-6)


def test_rating_curve_is_not_a_power_law() -> None:
    """The literal power-law reading of the email summary is dead (ADR-0019 §2).

    At the Obihiro anchor the power law a*Q**b gives ~1e-115 m — about 120
    orders of magnitude below the correct Eq. 4.19 stage of ~37.95 m MSL. An
    implementation that regressed to the power law fails here unmistakably.
    """
    stage = float(apply_rating_curve(_OBIHIRO_PEAK_Q_M3S, _OBIHIRO_A_KP, _OBIHIRO_B_KP))
    assert stage == pytest.approx(37.95, abs=5e-3)
    power_law_reading = _OBIHIRO_A_KP * _OBIHIRO_PEAK_Q_M3S**_OBIHIRO_B_KP
    assert power_law_reading < 1e-100


def test_rating_curve_is_elementwise_on_arrays() -> None:
    """The conversion broadcasts elementwise over a discharge array."""
    discharge = np.array([4.0, 16.0, 36.0, 64.0])
    # sqrt(Q/4) + 10 -> [11, 12, 13, 14]
    expected = np.array([11.0, 12.0, 13.0, 14.0])
    result = apply_rating_curve(discharge, 4.0, -10.0)
    np.testing.assert_allclose(result, expected)


def test_rating_curve_rejects_negative_radicand() -> None:
    """A negative ``Q/a_kp`` is rejected before the square root (ADR-0019 §2).

    The guard exists to catch bad inputs (physical Q >= 0 with a_kp > 0 can
    never trip it). Pinned as ``ValueError`` per the module's validation
    convention rather than a bare ``assert`` (which ``python -O`` would strip).
    """
    with pytest.raises(ValueError):
        apply_rating_curve(-1.0, _A_KP, _B_KP)
    with pytest.raises(ValueError):
        apply_rating_curve(np.array([10.0, -0.5, 20.0]), _A_KP, _B_KP)


# --- Obihiro anchor (ADR-0019 §4) ---------------------------------------------
def test_obihiro_anchor_reproduces_2016_peak() -> None:
    """KP 56.6 coefficients reproduce the 2016 record peak just below HWL.

    ADR-0019 §4: with a = 140.33 and b = -32.49, a peak discharge of ~4,180
    m^3/s gives ~37.95 m MSL, which is the Chapter 3 statement that the 2016
    record peak came 0.19 m below the 38.14 m MSL design HWL. Stage and HWL
    share the MSL datum (§3), so the comparison needs no reconciliation.
    """
    peak_stage = float(
        apply_rating_curve(_OBIHIRO_PEAK_Q_M3S, _OBIHIRO_A_KP, _OBIHIRO_B_KP)
    )
    assert peak_stage == pytest.approx(37.95, abs=5e-3)
    assert peak_stage < _OBIHIRO_HWL_M_MSL
    assert _OBIHIRO_HWL_M_MSL - peak_stage == pytest.approx(0.19, abs=5e-3)


def test_obihiro_realistic_flood_range() -> None:
    """Realistic flood peaks of 3,000-7,500 m^3/s span 37.1-39.8 m MSL.

    The ADR-0019 §4 plausibility band around the anchor: the low end sits
    ~1 m below HWL and the high end ~1.7 m above it.
    """
    low = float(apply_rating_curve(3000.0, _OBIHIRO_A_KP, _OBIHIRO_B_KP))
    high = float(apply_rating_curve(7500.0, _OBIHIRO_A_KP, _OBIHIRO_B_KP))
    assert low == pytest.approx(37.11, abs=0.01)
    assert high == pytest.approx(39.80, abs=0.01)


# --- Rating-coefficient CSVs (ADR-0019 §5) ------------------------------------
def test_rating_csv_loader_reads_shift_jis_full_width_header(
    tmp_path: Path,
) -> None:
    """The loader reads Shift-JIS CSVs with full-width ``HQ_a``/``HQ_b`` headers.

    Byte-exact hermetic replica of the real ``HQrelation_*Riv_2017.csv``
    layout (ADR-0019 §5): Shift-JIS encoding, header cells ``HQ_ａ`` /
    ``HQ_ｂ`` (FULL-WIDTH a and b, U+FF41/U+FF42), columns River, KP, a,
    b, CRLF line endings, and KP cells in both integer form (``3``) and
    decimal form (``56.6``). The distinct a/b magnitudes make a swapped
    column mapping fail loudly.
    """
    header = "River,KP,HQ_ａ,HQ_ｂ"
    rows = ["Tokachi,3,1341.3,0.51", "Tokachi,56.6,140.33,-32.49"]
    csv_path = tmp_path / "HQrelation_fixture.csv"
    csv_path.write_bytes(("\r\n".join([header, *rows]) + "\r\n").encode("shift_jis"))

    # The full-width header bytes (0x82 0x81 / 0x82 0x82) are invalid UTF-8,
    # so this fixture genuinely exercises the explicit-encoding requirement.
    with pytest.raises(UnicodeDecodeError):
        csv_path.read_bytes().decode("utf-8")

    coefficients = load_rating_coefficients(csv_path)
    assert coefficients[56.6] == pytest.approx((140.33, -32.49))
    assert coefficients[3.0] == pytest.approx((1341.3, 0.51))


@requires_rating_csvs
def test_real_tokachi_csv_yields_obihiro_coefficients() -> None:
    """The genuine Tokachi CSV maps KP 56.6 to (a, b) = (140.33, -32.49).

    End-to-end column-mapping guard on the real file: the correct coefficient
    reaches the correct KP, with a in column 3 and b in column 4.
    """
    coefficients = load_rating_coefficients(_TOKACHI_CSV)
    assert coefficients[56.6] == pytest.approx((_OBIHIRO_A_KP, _OBIHIRO_B_KP))


@requires_rating_csvs
def test_real_csvs_cover_all_study_nodes() -> None:
    """Rating coefficients exist for every study node (ADR-0019 §5, §7).

    Tokachi coverage must include the upper nodes KP 62.0-62.8, whose
    discharge is proxied from KP 61.8 but whose rating stays per-node; the
    Satsunai file must reach its KP 48 limit.
    """
    tokachi = load_rating_coefficients(_TOKACHI_CSV)
    study_kps = [56.6, 57.4, 58.8, 60.0, 61.8, 62.0, 62.2, 62.4, 62.6, 62.8]
    missing = [kp for kp in study_kps if kp not in tokachi]
    assert missing == []

    satsunai = load_rating_coefficients(_SATSUNAI_CSV)
    assert 48.0 in satsunai
    a_kp, b_kp = satsunai[48.0]
    assert np.isfinite(a_kp) and a_kp > 0.0
    assert np.isfinite(b_kp)


@requires_rating_csvs
def test_csv_to_stage_end_to_end_at_obihiro() -> None:
    """Real CSV coefficients feed Eq. 4.19 and land on the 2016 anchor.

    The full conversion chain of ADR-0019 §2-§5 on genuine data: read the
    Shift-JIS CSV, look up KP 56.6, convert the ~4,180 m^3/s peak, and land
    0.19 m below the 38.14 m MSL HWL.
    """
    a_kp, b_kp = load_rating_coefficients(_TOKACHI_CSV)[56.6]
    stage = float(apply_rating_curve(_OBIHIRO_PEAK_Q_M3S, a_kp, b_kp))
    assert stage == pytest.approx(37.95, abs=5e-3)
    assert stage < _OBIHIRO_HWL_M_MSL


# --- Units at the boundary: hours -> seconds (ADR-0019 §6) --------------------
def test_time_array_is_converted_to_seconds() -> None:
    """The record's time axis is SI seconds, converted from the source's hours.

    The M3 unit-boundary guard (analogue of the M4 m/s-vs-m/day trap): feed a
    time axis in hours and require the record to carry ``t`` in seconds. ADR-
    0019 §6 calls leaving the axis in hours out as the common bug to assert
    against.
    """
    time_hours = np.array([0.0, 1.0, 2.0, 3.0])
    discharge = np.array([4.0, 16.0, 36.0, 64.0])
    record = build_hydrograph_record(
        time_hours,
        discharge,
        a_kp=_A_KP,
        b_kp=_B_KP,
        scenario="historical",
        event_id="units_check",
    )
    np.testing.assert_allclose(record.t, time_hours * _SECONDS_PER_HOUR)
    # Sanity: the last sample is 3 h == 10800 s, not left as 3.
    assert record.t[-1] == pytest.approx(10800.0)


# --- native_dt is derived from the Time column, not hardcoded -----------------
def test_native_dt_reflects_hourly_source_resolution() -> None:
    """An hourly source yields native_dt == 3600 s (ADR-0019 §6)."""
    time_hours, discharge = _synthetic_hourly_discharge()
    record = build_hydrograph_record(
        time_hours,
        discharge,
        a_kp=_A_KP,
        b_kp=_B_KP,
        scenario="historical",
        event_id="synthetic_0001",
    )
    assert record.native_dt == pytest.approx(_SECONDS_PER_HOUR)


@pytest.mark.parametrize(
    ("spacing_hours", "expected_native_dt_s"),
    [
        (1.0, 3600.0),  # hourly (the confirmed, final d4PDF resolution)
        (0.5, 1800.0),  # 30-min: must NOT come back as 3600
        (2.0, 7200.0),  # 2-hourly: must NOT come back as 3600
    ],
)
def test_native_dt_is_derived_from_actual_spacing(
    spacing_hours: float, expected_native_dt_s: float
) -> None:
    """``native_dt`` is computed from the real spacing, not assumed 3600 s.

    ADR-0019 §6: native_dt = 3600 s is *derived from the Time column*, not
    hardcoded. The 0.5 h and 2.0 h cases fail against any hardcoded-hourly
    implementation.
    """
    time_hours = np.arange(6, dtype=np.float64) * spacing_hours
    discharge = np.linspace(10.0, 60.0, time_hours.size)
    record = build_hydrograph_record(
        time_hours,
        discharge,
        a_kp=_A_KP,
        b_kp=_B_KP,
        scenario="historical",
        event_id="spacing_check",
    )
    assert record.native_dt == pytest.approx(expected_native_dt_s)


def test_d4pdf_192_hour_event_window() -> None:
    """The real d4PDF Time column (integer hours 1..192) round-trips correctly.

    ADR-0019 §1: each sheet carries Time = 1..192 h (~8-day window). The
    record must span 191 elapsed hours at native_dt = 3600 s with the first
    sample at t = 3600 s (1 h), not renumbered to zero.
    """
    time_hours = np.arange(1, 193, dtype=np.float64)
    discharge = np.full(time_hours.size, 500.0)
    record = build_hydrograph_record(
        time_hours,
        discharge,
        a_kp=_A_KP,
        b_kp=_B_KP,
        scenario="historical",
        event_id="HPB_m001_1951",
    )
    assert record.t.size == 192
    assert record.t[0] == pytest.approx(3600.0)
    assert record.native_dt == pytest.approx(3600.0)
    assert record.duration_hours == pytest.approx(191.0)


def test_non_uniform_time_spacing_is_rejected() -> None:
    """A non-uniformly spaced time axis is rejected (native_dt ill-defined).

    ``native_dt`` is a single scalar resolution, so the loader must refuse a
    time axis whose spacing is not constant rather than silently pick one gap.
    """
    time_hours = np.array([0.0, 1.0, 2.0, 3.5])  # last gap is 1.5 h
    discharge = np.array([10.0, 20.0, 30.0, 40.0])
    with pytest.raises(ValueError):
        build_hydrograph_record(
            time_hours,
            discharge,
            a_kp=_A_KP,
            b_kp=_B_KP,
            scenario="historical",
            event_id="bad_spacing",
        )


# --- monotonic time ----------------------------------------------------------
@pytest.mark.parametrize(
    "time_hours",
    [
        np.array([0.0, 1.0, 1.0, 2.0]),  # repeated (non-strict)
        np.array([0.0, 1.0, 0.5, 2.0]),  # decreasing step
    ],
)
def test_time_axis_must_be_strictly_increasing(time_hours: np.ndarray) -> None:
    """A non-strictly-increasing time axis is rejected."""
    discharge = np.array([10.0, 20.0, 30.0, 40.0])
    with pytest.raises(ValueError):
        build_hydrograph_record(
            time_hours,
            discharge,
            a_kp=_A_KP,
            b_kp=_B_KP,
            scenario="historical",
            event_id="non_monotonic",
        )


# --- peak, duration, and the single conversion path --------------------------
def test_record_h_equals_rating_curve_of_discharge() -> None:
    """The record's h(t) is exactly Eq. 4.19 applied to Q(t).

    Ties record construction to the single conversion function, so a future
    rating-curve change propagates through the record with no separate code
    path.
    """
    time_hours, discharge = _synthetic_hourly_discharge()
    record = build_hydrograph_record(
        time_hours,
        discharge,
        a_kp=_A_KP,
        b_kp=_B_KP,
        scenario="historical",
        event_id="synthetic_0001",
    )
    np.testing.assert_allclose(record.h, apply_rating_curve(discharge, _A_KP, _B_KP))


def test_peak_is_max_stage() -> None:
    """``peak`` is the maximum stage, i.e. Eq. 4.19 at the peak discharge.

    Eq. 4.19 is monotonically increasing in Q (a_kp > 0), so the peak stage
    occurs at the peak discharge; here Q_max = 1600 with (a, b) = (4, -30)
    gives peak = sqrt(1600/4) + 30 = 50.0.
    """
    time_hours, discharge = _synthetic_hourly_discharge()
    record = build_hydrograph_record(
        time_hours,
        discharge,
        a_kp=_A_KP,
        b_kp=_B_KP,
        scenario="historical",
        event_id="synthetic_0001",
    )
    assert record.peak == pytest.approx(50.0)
    assert record.peak == pytest.approx(float(record.h.max()))


def test_duration_hours_is_elapsed_span() -> None:
    """``duration_hours`` is the elapsed span in hours.

    25 hourly samples span 24 hours.
    """
    time_hours, discharge = _synthetic_hourly_discharge()
    record = build_hydrograph_record(
        time_hours,
        discharge,
        a_kp=_A_KP,
        b_kp=_B_KP,
        scenario="historical",
        event_id="synthetic_0001",
    )
    assert record.duration_hours == pytest.approx(24.0)


# --- no crest capping inside M3 (ADR-0019 §8) ---------------------------------
def test_stage_is_not_capped_at_crest() -> None:
    """Stage exceeding the design HWL passes through M3 uncapped.

    ADR-0019 §8: Uemura's failure model replaces stage with crest height when
    stage exceeds the crest, but that capping belongs to the overflow
    mechanism, NOT to the seepage boundary condition — the BEP engine consumes
    the full uncapped stage. With the Obihiro coefficients, Q = 7,500 m^3/s
    gives 39.80 m MSL, 1.66 m above the 38.14 m HWL, and must survive intact.
    """
    time_hours = np.arange(1.0, 6.0)
    discharge = np.array([3000.0, 4180.0, 7500.0, 4180.0, 3000.0])
    record = build_hydrograph_record(
        time_hours,
        discharge,
        a_kp=_OBIHIRO_A_KP,
        b_kp=_OBIHIRO_B_KP,
        scenario="+4K",
        event_id="crest_exceedance",
    )
    assert record.peak == pytest.approx(39.80, abs=5e-3)
    assert record.peak > _OBIHIRO_HWL_M_MSL
    # Every sample equals the raw conversion — nothing was clipped to HWL.
    np.testing.assert_allclose(
        record.h, apply_rating_curve(discharge, _OBIHIRO_A_KP, _OBIHIRO_B_KP)
    )
    assert np.count_nonzero(record.h > _OBIHIRO_HWL_M_MSL) == 1


# --- scenario tags (ADR-0019 §9) -----------------------------------------------
@pytest.mark.parametrize("scenario", ["historical", "+4K"])
def test_scenario_tag_is_carried_through(scenario: str) -> None:
    """The scenario tag ('historical' or '+4K') round-trips onto the record.

    The scenario field flows through to ``FragilityResult`` metadata and the
    climate comparison depends on it, so both tags are exercised.
    """
    time_hours, discharge = _synthetic_hourly_discharge()
    record = build_hydrograph_record(
        time_hours,
        discharge,
        a_kp=_A_KP,
        b_kp=_B_KP,
        scenario=scenario,
        event_id="synthetic_0001",
    )
    assert record.scenario == scenario


# --- member column headers (ADR-0019 §1, §9) -----------------------------------
def test_parse_member_header_hpb() -> None:
    """``HPB_mXXX_YYYY`` parses to historical provenance with no SST pattern.

    The past experiment has a 50-flat-member structure: member ID and calendar
    year are preserved verbatim, ``sst`` is None, and HPB maps to the
    ``'historical'`` scenario tag.
    """
    info = parse_member_header("HPB_m001_1951")
    assert info["experiment"] == "HPB"
    assert info["scenario"] == "historical"
    assert info["sst"] is None
    assert info["member_id"] == "m001"
    assert info["year"] == 1951


@pytest.mark.parametrize("sst", _SST_PATTERNS)
def test_parse_member_header_hfb_all_sst_patterns(sst: str) -> None:
    """``HFB_{SST}_mXXX_YYYY`` parses to +4K provenance preserving the SST.

    The future experiment carries six prescribed sea-surface-temperature
    patterns (CC, GF, HA, MI, MP, MR; 15 members each); the pattern must
    survive into provenance so the 50-flat-past versus six-SST-future
    structure stays traceable (ADR-0019 §1).
    """
    info = parse_member_header(f"HFB_{sst}_m101_2051")
    assert info["experiment"] == "HFB"
    assert info["scenario"] == "+4K"
    assert info["sst"] == sst
    assert info["member_id"] == "m101"
    assert info["year"] == 2051


@pytest.mark.parametrize(
    "header",
    [
        "Time",  # the time column is not a member header
        "HPB_m001",  # missing year
        "HFB_m001_2051",  # HFB without an SST pattern
        "HFB_ZZ_m001_2051",  # unknown SST pattern
        "HXB_m001_1951",  # unknown experiment
    ],
)
def test_parse_member_header_rejects_invalid(header: str) -> None:
    """Non-member and malformed headers are rejected with ``ValueError``."""
    with pytest.raises(ValueError):
        parse_member_header(header)


# --- provenance on the record (ADR-0019 §1, §7) ---------------------------------
def test_provenance_defaults_to_empty_dict() -> None:
    """Without explicit provenance the record carries an empty dict.

    Keeps the same interface usable for the 2016 observed hydrograph, which
    has no ensemble member, year, or SST pattern (ADR-0019 Consequences).
    """
    time_hours, discharge = _synthetic_hourly_discharge()
    record = build_hydrograph_record(
        time_hours,
        discharge,
        a_kp=_A_KP,
        b_kp=_B_KP,
        scenario="historical",
        event_id="obs_2016",
    )
    assert record.provenance == {}


def test_provenance_flows_from_header_to_record() -> None:
    """Header provenance plus KP and proxy flags round-trip onto the record.

    Integration of ADR-0019 §1 (header provenance), §7 (upper-Tokachi
    discharge proxy: KP 62.0 uses the KP 61.8 series under its OWN local
    rating) and §9 (scenario tag): a +4K member's SST/member/year, the node
    KP, and ``discharge_proxied_from`` all survive onto the record.
    """
    header = "HFB_MR_m105_2062"
    info = parse_member_header(header)
    time_hours, discharge = _synthetic_hourly_discharge()
    record = build_hydrograph_record(
        time_hours,
        discharge,
        a_kp=_A_KP,
        b_kp=_B_KP,
        scenario=info["scenario"],
        event_id=header,
        provenance={**info, "kp": 62.0, "discharge_proxied_from": "KP61.8"},
    )
    assert record.scenario == "+4K"
    assert record.provenance["sst"] == "MR"
    assert record.provenance["member_id"] == "m105"
    assert record.provenance["year"] == 2062
    assert record.provenance["kp"] == pytest.approx(62.0)
    assert record.provenance["discharge_proxied_from"] == "KP61.8"


# --- Real d4PDF workbook seam (opt-in smoke tests; ADR-0019 §1, §6, §7) ---------
# These are the only tests that open the genuine multi-thousand-column Excel
# files; everything above locks the pure layer hermetically. They are marked
# ``slow`` (a full read-only streaming pass is ~5 s per workbook) and skip when
# the untracked data drop is absent.


@requires_hpb_workbook
@pytest.mark.slow
def test_real_hpb_workbook_loads_full_ensemble_at_kp56_6() -> None:
    """The genuine HPB band workbook loads end-to-end into 3,000 stage records.

    The full ADR-0019 chain on real data: openpyxl read-only streaming of the
    ``QT`` sheet (§1), header provenance (§1/§9), hourly Time -> SI seconds
    with derived ``native_dt`` = 3600 s (§6), and Eq. 4.19 under the node's
    own rating (§2/§5). Landing in the Obihiro MSL band (~33-40 m, §4) is the
    smoke check that the datum term ``-b_kp`` reached the stages.
    """
    rating = load_rating_coefficients(_TOKACHI_CSV)
    records = load_hydrograph_ensemble(
        _HPB_UPPER_TOKACHI_XLSX, kp=56.6, rating_coefficients=rating
    )

    # 3,000 columns: the complete past-experiment ensemble (ADR-0019 §1).
    assert len(records) == 3000

    sample = records["HPB_m001_1951"]
    assert isinstance(sample, HydrographRecord)
    assert sample.native_dt == 3600.0  # derived from the hourly Time column
    assert sample.t[0] == 3600.0  # Time = 1..192 h, converted, not renumbered
    assert sample.duration_hours == pytest.approx(191.0)
    assert sample.scenario == "historical"
    assert sample.provenance["experiment"] == "HPB"
    assert sample.provenance["member_id"] == "m001"
    assert sample.provenance["year"] == 1951
    assert sample.provenance["sst"] is None
    assert sample.provenance["kp"] == pytest.approx(56.6)
    assert "discharge_proxied_from" not in sample.provenance  # own coverage

    # Every record's stage lies in the plausible Obihiro MSL band (ADR-0019
    # §4: realistic peaks span 37.1-39.8 m MSL; base flow sits around ~33 m).
    # A stage near 0 m would mean b_kp never entered; near 75 m would mean raw
    # discharge leaked through as stage.
    peaks = np.array([r.peak for r in records.values()])
    assert peaks.min() > 30.0
    assert peaks.max() < 42.0
    assert all(r.peak == r.h.max() for r in list(records.values())[:10])


@requires_hpb_workbook
@pytest.mark.slow
def test_real_hpb_workbook_proxy_path_at_kp62_0() -> None:
    """KP 62.0 loads the KP 61.8 band workbook under its OWN local rating (§7).

    The upper-Tokachi coverage-gap rule on real data: the returned records are
    stamped ``discharge_proxied_from = 'KP61.8'`` and their stages are the
    KP 62.0 rating applied to the band discharge — verified by rebuilding one
    member's stage from the raw workbook column with the KP 62.0 coefficients
    and requiring an exact match (and a mismatch under the KP 61.8 rating).
    """
    rating = load_rating_coefficients(_TOKACHI_CSV)
    records = load_hydrograph_ensemble(
        _HPB_UPPER_TOKACHI_XLSX, kp=62.0, rating_coefficients=rating
    )
    assert len(records) == 3000

    sample = records["HPB_m001_1951"]
    assert sample.provenance["kp"] == pytest.approx(62.0)
    assert sample.provenance["discharge_proxied_from"] == "KP61.8"

    _, members = read_discharge_ensemble(_HPB_UPPER_TOKACHI_XLSX)
    q = members["HPB_m001_1951"]
    a_62_0, b_62_0 = rating[62.0]
    a_61_8, b_61_8 = rating[61.8]
    np.testing.assert_array_equal(sample.h, apply_rating_curve(q, a_62_0, b_62_0))
    assert not np.allclose(sample.h, apply_rating_curve(q, a_61_8, b_61_8))


@requires_hfb_workbook
@pytest.mark.slow
def test_real_hfb_workbook_headers_parse_to_plus4k() -> None:
    """The genuine HFB workbook's member headers all parse to '+4K' (§1, §9).

    Header-layer smoke test on the future-experiment file: 5,400 member
    columns, each ``HFB_{SST}_mXXX_YYYY`` with a known SST pattern, so the
    scenario tag that drives the climate comparison is producible from the
    real data. Reads only the header row (the discharge body is exercised on
    the HPB file above).
    """
    from openpyxl import load_workbook

    workbook = load_workbook(_HFB_UPPER_TOKACHI_XLSX, read_only=True, data_only=True)
    try:
        header_row = next(workbook["QT"].iter_rows(values_only=True))
    finally:
        workbook.close()

    member_headers = [str(c) for c in header_row[1:] if c is not None]
    assert len(member_headers) == 5400
    parsed = [parse_member_header(h) for h in member_headers]
    assert {p["scenario"] for p in parsed} == {"+4K"}
    assert {p["sst"] for p in parsed} == set(_SST_PATTERNS)


# --- MSL-datum guard for the M8 feed (ADR-0018/0019 §3; audit gap G2) -----------
def test_datum_guard_refuses_provisional_z_toe() -> None:
    """An MSL-stage record is refused against the provisional z_toe = 0.0.

    M3 stages are m MSL by construction (ADR-0019 §3) while the generated
    configs still carry the PROVISIONAL ``z_toe = 0.0`` (ADR-0018): pairing
    the two would silently hand M8 ~35 m driving heads. The guard makes that
    incompatibility loud — it must raise before any real-hydrograph record
    reaches the M8 fragility path with an unresolved datum.
    """
    time_hours = np.array([1.0, 2.0, 3.0, 4.0])
    record = build_hydrograph_record(
        time_hours,
        np.array([1000.0, 4180.0, 2000.0, 1000.0]),
        a_kp=_OBIHIRO_A_KP,
        b_kp=_OBIHIRO_B_KP,
        scenario="historical",
        event_id="datum_guard_case",
    )

    with pytest.raises(ValueError, match="z_toe"):
        validate_datum_consistency(record, z_toe_m=0.0)

    # A true MSL toe elevation passes silently (returns None, raises nothing).
    assert validate_datum_consistency(record, z_toe_m=33.0) is None


# --- KP -> band-file resolution + scenario mapping (ADR-0020) -------------------
_BAND_FILENAMES = [
    "Hydro Data, HPB, Tokachi Riv. KP053.40-KP056.00.xlsx",
    "Hydro Data, HPB, Tokachi Riv. KP056.20-KP061.80.xlsx",
    "Hydro Data, HFB, Tokachi Riv. KP056.20-KP061.80.xlsx",
    "Hydro Data, HPB, Satsunai Riv. KP002.80-KP024.60.xlsx",
]


@pytest.fixture()
def fake_data_root(tmp_path: Path) -> Path:
    """A hermetic data_root with the real drop's band-workbook names (empty)."""
    hydro_dir = tmp_path / "hydrographs"
    hydro_dir.mkdir()
    for name in _BAND_FILENAMES:
        (hydro_dir / name).touch()
    return tmp_path


def test_experiment_for_scenario_mapping() -> None:
    """historical -> HPB, +4K -> HFB; anything else is rejected (ADR-0020 §3)."""
    assert experiment_for_scenario("historical") == "HPB"
    assert experiment_for_scenario("+4K") == "HFB"
    with pytest.raises(ValueError):
        experiment_for_scenario("rcp85")


def test_rating_curve_path_convention() -> None:
    """The rating CSV path derives from the river name (ADR-0020 §2)."""
    path = rating_curve_path("data/raw", "Tokachi")
    assert path == Path("data/raw") / "rating_curves" / "HQrelation_TokachiRiv_2017.csv"
    assert rating_curve_path("x", "Satsunai").name == "HQrelation_SatsunaiRiv_2017.csv"
    with pytest.raises(ValueError):
        rating_curve_path("data/raw", "Ishikari")


@pytest.mark.parametrize(
    ("kp", "scenario", "expected"),
    [
        # In-band nodes select their own band under the scenario's experiment.
        (57.4, "historical", "Hydro Data, HPB, Tokachi Riv. KP056.20-KP061.80.xlsx"),
        (57.4, "+4K", "Hydro Data, HFB, Tokachi Riv. KP056.20-KP061.80.xlsx"),
        (54.0, "historical", "Hydro Data, HPB, Tokachi Riv. KP053.40-KP056.00.xlsx"),
        # Band edges are inclusive.
        (56.2, "historical", "Hydro Data, HPB, Tokachi Riv. KP056.20-KP061.80.xlsx"),
        (61.8, "historical", "Hydro Data, HPB, Tokachi Riv. KP056.20-KP061.80.xlsx"),
        # ADR-0019 §7 proxy: KP 62.0-62.8 route to the KP 61.8 band file.
        (62.0, "historical", "Hydro Data, HPB, Tokachi Riv. KP056.20-KP061.80.xlsx"),
        (62.8, "historical", "Hydro Data, HPB, Tokachi Riv. KP056.20-KP061.80.xlsx"),
    ],
)
def test_resolve_band_workbook_selects_the_right_file(
    fake_data_root: Path, kp: float, scenario: str, expected: str
) -> None:
    """The resolver picks the unique band file for (river, kp, scenario).

    Covers the ADR-0020 §4 rules: filename-parsed bands (not hardcoded),
    scenario -> experiment selection, inclusive band edges, and the §7
    upper-Tokachi proxy routing through ``resolve_discharge_source_kp``.
    """
    path = resolve_band_workbook(
        fake_data_root, river="Tokachi", kp=kp, scenario=scenario
    )
    assert path.name == expected


def test_resolve_band_workbook_selects_by_river(fake_data_root: Path) -> None:
    """A Satsunai node never resolves to a Tokachi band file."""
    path = resolve_band_workbook(
        fake_data_root, river="Satsunai", kp=10.0, scenario="historical"
    )
    assert path.name == "Hydro Data, HPB, Satsunai Riv. KP002.80-KP024.60.xlsx"


def test_resolve_band_workbook_no_match_is_loud(fake_data_root: Path) -> None:
    """Zero matching band files raise ValueError (no silent nearest-band)."""
    with pytest.raises(ValueError, match="[Nn]o .*band"):
        resolve_band_workbook(
            fake_data_root, river="Tokachi", kp=70.0, scenario="historical"
        )
    # A scenario whose experiment file set is absent is equally loud: the
    # fake drop has no HFB Satsunai workbook.
    with pytest.raises(ValueError):
        resolve_band_workbook(fake_data_root, river="Satsunai", kp=10.0, scenario="+4K")


def test_resolve_band_workbook_ambiguous_is_loud(fake_data_root: Path) -> None:
    """Two bands covering the same KP raise ValueError naming both files."""
    overlap = "Hydro Data, HPB, Tokachi Riv. KP056.00-KP062.00.xlsx"
    (fake_data_root / "hydrographs" / overlap).touch()
    with pytest.raises(ValueError, match="ambiguous"):
        resolve_band_workbook(
            fake_data_root, river="Tokachi", kp=57.4, scenario="historical"
        )


def test_resolve_band_workbook_missing_dir_is_loud(tmp_path: Path) -> None:
    """A data_root without a hydrographs/ directory raises ValueError."""
    with pytest.raises(ValueError, match="hydrographs"):
        resolve_band_workbook(tmp_path, river="Tokachi", kp=57.4, scenario="historical")


@requires_hpb_workbook
def test_resolve_band_workbook_on_real_data_drop() -> None:
    """Against the genuine data/raw drop, KP 57.4 resolves to the real file.

    Name-only check (no workbook read): the resolver, pointed at the actual
    untracked data drop, lands on the same file the smoke tests open.
    """
    data_root = Path(__file__).resolve().parents[1] / "data" / "raw"
    path = resolve_band_workbook(
        data_root, river="Tokachi", kp=57.4, scenario="historical"
    )
    assert path == _HPB_UPPER_TOKACHI_XLSX
    assert path.exists()


# --- Canonical shape: normalize + per-level scaling (G1; ADR-0020 Decision 1) ---
def _canonical_from_stage(
    h: np.ndarray, dt_s: float = 3600.0, event_id: str = "HPB_m064_1987"
) -> CanonicalShape:
    """Build a CanonicalShape directly from an in-memory stage series."""
    shape, h_base, _ = normalize_stage_shape(h)
    record = HydrographRecord(
        t=np.arange(1.0, h.size + 1.0) * dt_s,
        h=h,
        peak=float(h.max()),
        duration_hours=float((h.size - 1) * dt_s / 3600.0),
        scenario="historical",
        event_id=event_id,
        native_dt=dt_s,
        provenance={"experiment": "HPB", "member_id": "m064", "year": 1987},
    )
    return CanonicalShape(source_record=record, shape=shape, h_base_m=h_base)


def test_normalize_stage_shape_exact_endpoints() -> None:
    """The normalized shape hits 0.0 and 1.0 exactly at floor and peak.

    Exactness (not approx) is load-bearing: min(shape) == 0 pins the rescaled
    trough floor at h_base exactly, and max(shape) == 1 puts the rescaled
    peak at h_i to within one ulp.
    """
    h = np.array([32.0, 34.7, 37.3, 33.1, 36.0, 32.0])
    shape, h_base, h_peak = normalize_stage_shape(h)
    assert h_base == 32.0
    assert h_peak == 37.3
    assert shape.min() == 0.0
    assert shape.max() == 1.0
    # Stage-domain normalization: relative structure preserved linearly.
    np.testing.assert_allclose(shape, (h - 32.0) / (37.3 - 32.0), rtol=0, atol=0)

    with pytest.raises(ValueError, match="degenerate"):
        normalize_stage_shape(np.full(5, 33.0))


def test_conditioning_record_scales_to_level_with_h_base_floor() -> None:
    """The G1 scaling rule: trough pinned at h_base, peak = h_i verbatim.

    ``h(t) = h_base + (h_i - h_base) * shape(t)`` (ADR-0020 Decision 1;
    ADR-0021 item 4): the floor stays the base-flow stage — NOT z_toe — so
    inter-peak recessions are not artificially deepened, and the ``peak``
    field is the authoritative conditioning anchor set to h_i verbatim
    (ADR-0010; M8 ambiguity 3).
    """
    # Compound (two-peak) source: base 32, precursor 36, trough 33, main 38.
    h_src = np.array([32.0, 36.0, 33.0, 38.0, 32.0])
    canonical = _canonical_from_stage(h_src)
    assert canonical.h_base_m == 32.0

    level = 41.5
    record = conditioning_record_for_level(canonical, level, scenario="+4K")

    assert isinstance(record, HydrographRecord)
    assert record.peak == level  # verbatim anchor
    assert float(record.h.max()) == pytest.approx(level, abs=1e-12)
    assert float(record.h.min()) == canonical.h_base_m  # floor pinned exactly
    # Linear stage-domain rescaling preserves the event's relative structure:
    # the trough sits at h_base + (h_i - h_base) * (33-32)/(38-32).
    expected = 32.0 + (level - 32.0) * (h_src - 32.0) / 6.0
    np.testing.assert_allclose(record.h, expected, rtol=1e-15)

    # Time axis untouched (no time rescaling): full source window, native dt.
    np.testing.assert_array_equal(record.t, canonical.source_record.t)
    assert record.native_dt == canonical.source_record.native_dt
    assert record.duration_hours == canonical.source_record.duration_hours

    # Run scenario on the record; shape source stays in provenance.
    assert record.scenario == "+4K"
    assert record.provenance["shape_source_event"] == "HPB_m064_1987"
    assert record.provenance["conditioning_level_m_msl"] == level
    assert record.provenance["h_base_m_msl"] == 32.0
    assert record.provenance["experiment"] == "HPB"


def test_conditioning_record_below_h_base_is_constant_stage() -> None:
    """A level at or below h_base emits a constant-stage record at h_i.

    Sub-base-flow conditioning levels (the grids' lowest anchors sit up to
    0.04 m below h_base) have no positive amplitude to scale; a constant
    stage at h_i is the zero-load floor, not an inverted event.
    """
    canonical = _canonical_from_stage(np.array([32.0, 36.0, 33.0, 38.0, 32.0]))
    for level in (31.0, 32.0):
        record = conditioning_record_for_level(canonical, level, scenario="historical")
        np.testing.assert_array_equal(record.h, np.full(5, level))
        assert record.peak == level


def test_conditioning_record_is_deterministic() -> None:
    """Two calls with the same inputs yield bit-identical records (gap G4).

    The pure-function property the parallel == serial guarantee rests on.
    """
    canonical = _canonical_from_stage(np.array([32.0, 36.0, 33.0, 38.0, 32.0]))
    a = conditioning_record_for_level(canonical, 40.0, scenario="historical")
    b = conditioning_record_for_level(canonical, 40.0, scenario="historical")
    np.testing.assert_array_equal(a.h, b.h)
    assert a.peak == b.peak and a.event_id == b.event_id
    assert a.provenance == b.provenance


# --- load_canonical_shape: hermetic end-to-end on a fake data drop --------------
def _write_fake_drop(
    root: Path,
    *,
    kp_rows: list[tuple[float, float, float]],
    members: dict[str, list[float]],
    time_hours: list[float] | None = None,
    band: str = "KP056.20-KP061.80",
) -> None:
    """Write a minimal ADR-0020-layout data drop: rating CSV + band workbook."""
    from openpyxl import Workbook

    rating_dir = root / "rating_curves"
    rating_dir.mkdir(parents=True, exist_ok=True)
    lines = ["River,KP,HQ_ａ,HQ_ｂ"]  # full-width a/b, as on disk (ADR-0019 §5)
    lines += [f"Tokachi,{kp:g},{a:g},{b:g}" for kp, a, b in kp_rows]
    (rating_dir / "HQrelation_TokachiRiv_2017.csv").write_bytes(
        ("\r\n".join(lines) + "\r\n").encode("shift_jis")
    )

    hydro_dir = root / "hydrographs"
    hydro_dir.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "QT"
    names = list(members)
    sheet.append(["Time", *names])
    n_rows = len(next(iter(members.values())))
    hours = time_hours if time_hours is not None else list(range(1, n_rows + 1))
    for row_index in range(n_rows):
        sheet.append([hours[row_index], *(members[n][row_index] for n in names)])
    workbook.save(hydro_dir / f"Hydro Data, HPB, Tokachi Riv. {band}.xlsx")


def test_load_canonical_shape_hermetic_end_to_end(tmp_path: Path) -> None:
    """The composed loader builds the right shape from a fake data drop.

    Locks the composition (ADR-0020): rating convention -> band resolution by
    the EVENT's experiment -> member extraction -> stage under the node's own
    rating -> stage-domain normalization, with provenance carrying the member
    info and the resolved file names. Rating a=1, b=-30 makes stages exact:
    h = sqrt(Q) + 30.
    """
    # Q = [4, 49, 9, 100, 4] -> h = [32, 37, 33, 40, 32]: compound two-peak.
    _write_fake_drop(
        tmp_path,
        kp_rows=[(57.4, 1.0, -30.0)],
        members={
            "HPB_m064_1987": [4.0, 49.0, 9.0, 100.0, 4.0],
            "HPB_m067_1978": [4.0, 4.0, 100.0, 9.0, 4.0],
        },
    )
    canonical = load_canonical_shape(
        tmp_path, river="Tokachi", kp=57.4, event_id="HPB_m064_1987"
    )

    assert canonical.h_base_m == pytest.approx(32.0)
    np.testing.assert_allclose(
        canonical.source_record.h, [32.0, 37.0, 33.0, 40.0, 32.0]
    )
    np.testing.assert_allclose(canonical.shape, [0.0, 5 / 8, 1 / 8, 1.0, 0.0])
    assert canonical.source_record.native_dt == 3600.0
    prov = canonical.source_record.provenance
    assert prov["experiment"] == "HPB"
    assert prov["member_id"] == "m064"
    assert prov["year"] == 1987
    assert prov["kp"] == pytest.approx(57.4)
    assert "discharge_proxied_from" not in prov
    assert prov["band_workbook"].endswith(".xlsx")
    assert prov["rating_csv"] == "HQrelation_TokachiRiv_2017.csv"

    # A missing member is loud, naming the workbook.
    with pytest.raises(ValueError, match="HPB_m999_2000"):
        load_canonical_shape(
            tmp_path, river="Tokachi", kp=57.4, event_id="HPB_m999_2000"
        )


def test_load_canonical_shape_proxy_node_uses_local_rating(tmp_path: Path) -> None:
    """KP 62.0: KP 61.8 band discharge under the KP 62.0 rating, marker set.

    The ADR-0019 §7 rule through the composed loader: distinct a/b at 61.8
    vs 62.0 prove the LOCAL rating was applied to the proxied discharge.
    """
    _write_fake_drop(
        tmp_path,
        kp_rows=[(61.8, 1.0, -30.0), (62.0, 4.0, -40.0)],
        members={"HPB_m064_1987": [4.0, 49.0, 9.0, 100.0, 4.0]},
    )
    canonical = load_canonical_shape(
        tmp_path, river="Tokachi", kp=62.0, event_id="HPB_m064_1987"
    )
    # h = sqrt(Q/4) + 40 (the KP 62.0 rating), NOT sqrt(Q) + 30 (KP 61.8's).
    np.testing.assert_allclose(
        canonical.source_record.h, [41.0, 43.5, 41.5, 45.0, 41.0]
    )
    assert canonical.source_record.provenance["discharge_proxied_from"] == "KP61.8"
    assert canonical.h_base_m == pytest.approx(41.0)


# ---------------------------------------------------------------------------
# resample_record (ADR-0013 record-construction hook; ADR-0030 integration-dt
# policy). The refinement must change ONLY the integration grid: every native
# sample stays a node of the refined grid (bit-exact), interior samples are the
# linear interpolant, and peak/duration/identity are untouched.
# ---------------------------------------------------------------------------


def _single_peak_record() -> HydrographRecord:
    """A small uniform hourly record with an interior peak (m MSL)."""
    t = np.arange(5, dtype=np.float64) * 3600.0
    h = np.array([40.0, 42.0, 45.0, 41.0, 40.0])
    return HydrographRecord(
        t=t,
        h=h,
        peak=45.0,
        duration_hours=4.0,
        scenario="historical",
        event_id="resample_fixture",
        native_dt=3600.0,
        provenance={"member_id": "m000"},
    )


def test_resample_record_nested_grid_preserves_source_nodes() -> None:
    """Factor-4 refinement: native nodes bit-exact, interiors linear."""
    record = _single_peak_record()
    refined = resample_record(record, 900.0)

    assert refined.native_dt == 900.0
    assert refined.t.size == (record.t.size - 1) * 4 + 1
    # Every native sample is a node of the refined grid, bit-exact -- the
    # loading signal is unchanged, only the integration grid is refined.
    np.testing.assert_array_equal(refined.h[::4], record.h)
    np.testing.assert_array_equal(refined.t[::4], record.t)
    # Interior points are the linear interpolant (no new extremes, no ringing).
    np.testing.assert_allclose(refined.h[1:4], [40.5, 41.0, 41.5])
    assert refined.h.max() == record.h.max()
    # Identity and span untouched; provenance records the refinement.
    assert refined.peak == record.peak
    assert refined.duration_hours == record.duration_hours
    assert refined.event_id == record.event_id
    assert refined.provenance["resampled_from_native_dt_s"] == 3600.0
    assert refined.provenance["resample_factor"] == 4
    assert refined.provenance["member_id"] == "m000"


def test_resample_record_factor_one_is_identity() -> None:
    """target == native returns the record itself (no copy, no markers)."""
    record = _single_peak_record()
    assert resample_record(record, 3600.0) is record


def test_resample_record_rejects_coarsening_and_non_divisors() -> None:
    """Only integer subdivisions of the native grid are legal (ADR-0013)."""
    record = _single_peak_record()
    with pytest.raises(ValueError, match="integer subdivision"):
        resample_record(record, 7200.0)  # coarsening
    with pytest.raises(ValueError, match="integer subdivision"):
        resample_record(record, 1000.0)  # non-nested grid
    with pytest.raises(ValueError, match="must be > 0"):
        resample_record(record, 0.0)
